import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from pathlib import Path
import plotly.express as px
import plotly.graph_objects as go
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

# =========================================================
# PAGE CONFIG
# =========================================================
st.set_page_config(
    page_title="ShelfIQ 911",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =========================================================
# BRANDING
# =========================================================
APP_TITLE = "ShelfIQ 911"
APP_SUBTITLE = "Retail Analytics, Shelf Optimization, and Sell-In Intelligence"
LOGO_PATH = "logo.png"

# =========================================================
# THEME / CSS
# =========================================================
st.markdown("""
<style>
:root {
    --bg-1: #08111f;
    --bg-2: #0d1b2a;
    --bg-3: #111f33;
    --panel: rgba(255,255,255,0.96);
    --panel-soft: rgba(255,255,255,0.88);
    --ink: #0f172a;
    --muted: #667085;
    --line: #dbe5ef;
    --white: #ffffff;
    --navy: #0f172a;
    --blue: #2563eb;
    --blue-2: #1d4ed8;
    --teal: #0f766e;
    --amber: #b45309;
    --rose: #b42318;
    --green: #166534;
    --shadow-lg: 0 20px 44px rgba(8, 17, 31, 0.16);
    --shadow-md: 0 12px 28px rgba(8, 17, 31, 0.10);
}
.stApp {
    background:
      radial-gradient(circle at top right, rgba(37,99,235,0.16), transparent 18%),
      radial-gradient(circle at top left, rgba(15,118,110,0.10), transparent 16%),
      linear-gradient(180deg, #091321 0%, #0f1b2d 20%, #e9eef5 20%, #eef3f7 100%);
}
[data-testid="stAppViewContainer"] {
    background:
      radial-gradient(circle at top right, rgba(37,99,235,0.16), transparent 18%),
      radial-gradient(circle at top left, rgba(15,118,110,0.10), transparent 16%),
      linear-gradient(180deg, #091321 0%, #0f1b2d 20%, #e9eef5 20%, #eef3f7 100%);
}
.block-container {
    padding-top: 0.8rem;
    padding-bottom: 1.8rem;
    max-width: 1580px;
}
.block-container {
    padding-top: 1.1rem;
    padding-bottom: 1.8rem;
    max-width: 1580px;
}
h1, h2, h3 { letter-spacing: -0.02em; }
section[data-testid="stSidebar"] {
    background:
      linear-gradient(180deg, rgba(7,15,28,0.98) 0%, rgba(13,27,42,0.98) 100%);
    border-right: 1px solid rgba(255,255,255,0.08);
}
section[data-testid="stSidebar"] * {
    color: #e5edf7 !important;
}
section[data-testid="stSidebar"] .stMarkdown p,
section[data-testid="stSidebar"] .stCaption,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] .st-emotion-cache-16txtl3 {
    color: #d7e3f4 !important;
}
section[data-testid="stSidebar"] .stRadio > div,
section[data-testid="stSidebar"] .stFileUploader,
section[data-testid="stSidebar"] [data-baseweb="select"],
section[data-testid="stSidebar"] .stMultiSelect,
section[data-testid="stSidebar"] .stTextInput > div > div,
section[data-testid="stSidebar"] .stNumberInput > div > div {
    background: rgba(255,255,255,0.06) !important;
    border: 1px solid rgba(255,255,255,0.12) !important;
    border-radius: 18px !important;
}
section[data-testid="stSidebar"] .stRadio label {
    background: rgba(255,255,255,0.04);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 14px;
    padding: 8px 10px;
}
section[data-testid="stSidebar"] .stButton > button,
section[data-testid="stSidebar"] .stDownloadButton > button {
    background: linear-gradient(180deg, #2b6ef3 0%, #1d4ed8 100%) !important;
    color: white !important;
    border: 0 !important;
}
div[data-testid="stTabs"] {
    margin-top: 0.25rem;
}
div[data-testid="stTabs"] button {
    border-radius: 999px;
    padding: 0.62rem 1.05rem;
    border: 1px solid #d9e2ec;
    background: linear-gradient(180deg, #ffffff 0%, #f6f9fc 100%);
    color: #334155;
    font-weight: 700;
    box-shadow: 0 8px 18px rgba(15,23,42,0.06);
}
div[data-testid="stTabs"] button:hover {
    border-color: #cbd8e6;
    background: white;
    color: #0f172a;
}
div[data-testid="stTabs"] button[aria-selected="true"] {
    background: linear-gradient(180deg, #0f172a 0%, #1d4ed8 100%);
    border-color: #1d4ed8;
    color: #ffffff;
    box-shadow: 0 10px 24px rgba(29,78,216,0.24);
}
.stButton > button, .stDownloadButton > button {
    border-radius: 15px !important;
    min-height: 47px;
    font-weight: 700;
    border: 1px solid #d6e0eb !important;
    box-shadow: 0 10px 22px rgba(8,17,31,0.08);
}
.hero-shell {
    background:
      radial-gradient(circle at top right, rgba(96,165,250,0.18), transparent 26%),
      radial-gradient(circle at left center, rgba(45,212,191,0.10), transparent 18%),
      linear-gradient(135deg, #0b1728 0%, #12233b 52%, #132948 100%);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 30px;
    padding: 26px 28px;
    box-shadow: 0 26px 46px rgba(8,17,31,0.22);
    margin-bottom: 0.9rem;
}
.hero-grid {
    display: grid;
    grid-template-columns: 90px 1.5fr 1fr;
    gap: 20px;
    align-items: center;
}
.hero-logo-wrap {
    width: 82px;
    height: 82px;
    border-radius: 22px;
    background: rgba(255,255,255,0.08);
    display:flex;
    align-items:center;
    justify-content:center;
    border: 1px solid rgba(255,255,255,0.08);
}
.hero-kicker {
    color: #8fb7ff;
    text-transform: uppercase;
    letter-spacing: 0.14em;
    font-size: 0.72rem;
    font-weight: 780;
    margin-bottom: 8px;
}
.hero-title {
    color: #ffffff;
    font-size: 2.15rem;
    line-height: 1.02;
    font-weight: 820;
    margin-bottom: 10px;
}
.hero-copy {
    color: #d7e4f6;
    font-size: 0.97rem;
    line-height: 1.58;
    max-width: 820px;
}
.hero-chip-row {
    display:flex;
    flex-wrap:wrap;
    gap:10px;
    margin-top: 14px;
}
.hero-chip {
    background: rgba(255,255,255,0.08);
    border: 1px solid rgba(255,255,255,0.09);
    color: #edf4ff;
    border-radius: 999px;
    padding: 8px 12px;
    font-size: 0.78rem;
    font-weight: 650;
}
.hero-panel {
    background: rgba(255,255,255,0.08);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 22px;
    padding: 16px 16px 14px 16px;
}
.hero-panel-title {
    color: #c9dbf7;
    font-size: 0.76rem;
    text-transform: uppercase;
    letter-spacing: 0.1em;
    margin-bottom: 10px;
    font-weight: 760;
}
.hero-panel-grid {
    display:grid;
    grid-template-columns: 1fr 1fr;
    gap: 10px;
}
.hero-stat {
    background: rgba(255,255,255,0.08);
    border-radius: 16px;
    padding: 10px 12px;
}
.hero-stat-label {
    color: #c9dbf7;
    font-size: 0.68rem;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 6px;
}
.hero-stat-value {
    color: white;
    font-size: 1rem;
    font-weight: 760;
}
.metric-card {
    background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%);
    border: 1px solid #dce6f0;
    border-radius: 24px;
    padding: 20px 20px 18px 20px;
    min-height: 128px;
    box-shadow: var(--shadow-md);
    position: relative;
    overflow: hidden;
}
.metric-card:before {
    content:"";
    position:absolute;
    left:0; top:0; bottom:0;
    width: 5px;
    background: linear-gradient(180deg, #2b6ef3 0%, #0f766e 100%);
}
.metric-label {
    color: #667085;
    font-size: 0.75rem;
    text-transform: uppercase;
    letter-spacing: 0.11em;
    margin-bottom: 10px;
    padding-left: 6px;
}
.metric-value {
    color: #101828;
    font-size: 1.95rem;
    line-height: 1.02;
    font-weight: 800;
    padding-left: 6px;
}
.metric-sub {
    color: #475467;
    font-size: 0.9rem;
    margin-top: 10px;
    padding-left: 6px;
}
.panel {
    background: var(--panel);
    border: 1px solid var(--line);
    border-radius: 24px;
    padding: 20px 20px 16px 20px;
    box-shadow: var(--shadow-md);
}
.section-title {
    font-size: 1rem;
    font-weight: 760;
    color: #0f172a;
    margin-bottom: 0.35rem;
}
.small-note {
    color: #667085;
    font-size: 0.9rem;
}
.executive-band {
    background: linear-gradient(135deg, #ffffff 0%, #f8fbff 54%, #eef5ff 100%);
    border: 1px solid #dce7f5;
    border-radius: 28px;
    padding: 22px 24px;
    box-shadow: var(--shadow-md);
    margin-bottom: 0.65rem;
}
.executive-kicker {
    color: #2563eb;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    font-size: 0.72rem;
    font-weight: 780;
    margin-bottom: 8px;
}
.executive-title {
    color: #0f172a;
    font-size: 1.9rem;
    font-weight: 820;
    line-height: 1.04;
    margin-bottom: 10px;
}
.executive-copy {
    color: #475467;
    font-size: 0.98rem;
    line-height: 1.58;
}
.signal-strip {
    display: grid;
    grid-template-columns: repeat(4, minmax(0,1fr));
    gap: 12px;
    margin-top: 14px;
}
.signal-card {
    background: rgba(255,255,255,0.88);
    border: 1px solid #dfe8f2;
    border-radius: 18px;
    padding: 12px 14px;
}
.signal-label {
    color: #667085;
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 6px;
}
.signal-value {
    color: #0f172a;
    font-size: 1rem;
    font-weight: 760;
}
.insight-card {
    background: linear-gradient(180deg, #ffffff 0%, #fbfcfe 100%);
    border: 1px solid #e2eaf2;
    border-radius: 24px;
    padding: 18px;
    box-shadow: var(--shadow-md);
    margin: 0.12rem 0 0.7rem 0;
}
.insight-header {
    display:flex;
    align-items:center;
    justify-content:space-between;
    gap:12px;
    margin-bottom: 12px;
}
.insight-kicker {
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.12em;
    color: #2563eb;
    font-weight: 780;
}
.insight-headline {
    font-size: 1.12rem;
    font-weight: 780;
    color: #0f172a;
    margin: 0;
}
.insight-sub {
    color: #667085;
    font-size: 0.88rem;
}
.insight-grid {
    display:grid;
    grid-template-columns: repeat(3, minmax(0,1fr));
    gap: 12px;
}
.insight-pill {
    background: linear-gradient(180deg, #f9fbfd 0%, #f5f8fc 100%);
    border: 1px solid #e4ebf2;
    border-radius: 18px;
    padding: 14px 14px 12px 14px;
}
.insight-pill-label {
    color: #667085;
    font-size: 0.72rem;
    font-weight: 720;
    text-transform: uppercase;
    letter-spacing: 0.08em;
    margin-bottom: 8px;
}
.insight-pill-value {
    color: #101828;
    font-size: 0.92rem;
    line-height: 1.5;
}
.ai-box {
    background:
      radial-gradient(circle at top right, rgba(96,165,250,0.18), transparent 26%),
      linear-gradient(180deg, #0f172a 0%, #15253d 100%);
    border-radius: 26px;
    border: 1px solid rgba(255,255,255,0.08);
    color: #e2e8f0;
    padding: 22px 22px 18px 22px;
    box-shadow: 0 18px 38px rgba(8,17,31,0.18);
}
.ai-box ul {
    margin: 0.6rem 0 0 1rem !important;
    padding-left: 1rem !important;
}
.ai-box li {
    color: #edf4ff !important;
    font-size: 0.96rem !important;
    line-height: 1.65 !important;
    margin-bottom: 0.6rem !important;
}
.ai-box p, .ai-box strong {
    color: #edf4ff !important;
}
.ai-box strong {
    font-size: 1rem !important;
    letter-spacing: 0.01em;
}
.badge {
    display:inline-flex;
    align-items:center;
    padding: 0.30rem 0.62rem;
    border-radius: 999px;
    font-size: 0.74rem;
    font-weight: 720;
    letter-spacing: 0.02em;
}
.badge-good { background: #ecfdf3; color: #166534; }
.badge-warn { background: #fff7ed; color: #b45309; }
.badge-risk { background: #fff1f2; color: #b42318; }
.download-panel {
    background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%);
    border: 1px solid #dce7f2;
    border-radius: 24px;
    padding: 16px;
    box-shadow: var(--shadow-md);
}
div[data-testid="stFileUploaderDropzone"] {
    background: rgba(255,255,255,0.04) !important;
    border: 1px dashed rgba(255,255,255,0.18) !important;
}
@media (max-width: 980px) {
    .hero-grid { grid-template-columns: 1fr; }
    .signal-strip, .insight-grid, .hero-panel-grid { grid-template-columns: 1fr; }
    .hero-title { font-size: 1.6rem; }
    .executive-title { font-size: 1.5rem; }
}

div[data-testid="stDataFrame"] {
    border: 1px solid #dbe5ef;
    border-radius: 20px;
    overflow: hidden;
    box-shadow: 0 10px 24px rgba(15,23,42,0.06);
    background: white;
}
div[data-testid="stMetric"] {
    background: transparent;
}

</style>
""", unsafe_allow_html=True)

# =========================================================
# STATE / REGION MAP
# =========================================================
STATE_TO_REGION = {
    "CT": "Northeast", "ME": "Northeast", "MA": "Northeast", "NH": "Northeast",
    "RI": "Northeast", "VT": "Northeast", "NJ": "Northeast", "NY": "Northeast",
    "PA": "Northeast",
    "IL": "Midwest", "IN": "Midwest", "MI": "Midwest", "OH": "Midwest",
    "WI": "Midwest", "IA": "Midwest", "KS": "Midwest", "MN": "Midwest",
    "MO": "Midwest", "NE": "Midwest", "ND": "Midwest", "SD": "Midwest",
    "AL": "Southeast", "AR": "Southeast", "DE": "Southeast", "DC": "Southeast",
    "FL": "Southeast", "GA": "Southeast", "KY": "Southeast", "LA": "Southeast",
    "MD": "Southeast", "MS": "Southeast", "NC": "Southeast", "SC": "Southeast",
    "TN": "Southeast", "VA": "Southeast", "WV": "Southeast",
    "AZ": "Southwest", "NM": "Southwest", "OK": "Southwest", "TX": "Southwest",
    "AK": "West", "CA": "West", "CO": "West", "HI": "West", "ID": "West",
    "MT": "West", "NV": "West", "OR": "West", "UT": "West", "WA": "West",
    "WY": "West"
}

# =========================================================
# LOGO HELPERS
# =========================================================
def get_logo_bytes(uploaded_logo=None):
    if uploaded_logo is not None:
        try:
            return uploaded_logo.getvalue()
        except Exception:
            try:
                uploaded_logo.seek(0)
                return uploaded_logo.read()
            except Exception:
                return None
    if Path(LOGO_PATH).exists():
        try:
            return Path(LOGO_PATH).read_bytes()
        except Exception:
            return None
    return None

def save_logo_bytes(logo_bytes):
    if logo_bytes:
        try:
            Path(LOGO_PATH).write_bytes(logo_bytes)
        except Exception:
            pass

def add_logo_to_sheet(ws, logo_bytes=None, cell="A1", width=135):
    if not logo_bytes:
        return
    try:
        bio = BytesIO(logo_bytes)
        img = XLImage(bio)
        ratio = img.height / max(img.width, 1)
        img.width = width
        img.height = max(int(width * ratio), 28)
        ws.add_image(img, cell)
    except Exception:
        pass

# =========================================================
# UI HELPERS
# =========================================================

def render_header(logo_bytes=None):
    logo_html = ""
    if logo_bytes:
        import base64
        encoded = base64.b64encode(logo_bytes).decode("utf-8")
        logo_html = f'<img src="data:image/png;base64,{encoded}" style="max-width:68px; max-height:68px; border-radius:14px;" />'
    elif Path(LOGO_PATH).exists():
        try:
            import base64
            encoded = base64.b64encode(Path(LOGO_PATH).read_bytes()).decode("utf-8")
            logo_html = f'<img src="data:image/png;base64,{encoded}" style="max-width:68px; max-height:68px; border-radius:14px;" />'
        except Exception:
            logo_html = '<div style="color:white;font-size:1.6rem;font-weight:800;">SI</div>'
    else:
        logo_html = '<div style="color:white;font-size:1.6rem;font-weight:800;">SI</div>'

    st.markdown(
        f"""
        <div class="hero-shell">
            <div class="hero-grid">
                <div class="hero-logo-wrap">{logo_html}</div>
                <div>
                    <div class="hero-kicker">Circana-style retail intelligence</div>
                    <div class="hero-title">{APP_TITLE}</div>
                    <div class="hero-copy">
                        Professional retail analytics with cleaner storytelling across distribution, velocity,
                        store productivity, quality risk, and growth signals. Built to feel like a modern
                        insights workspace instead of a basic demo dashboard.
                    </div>
                    <div class="hero-chip-row">
                        <div class="hero-chip">Executive KPI hierarchy</div>
                        <div class="hero-chip">Sharper trend visuals</div>
                        <div class="hero-chip">Cleaner risk indicators</div>
                        <div class="hero-chip">Action-based readouts</div>
                    </div>
                </div>
                <div class="hero-panel">
                    <div class="hero-panel-title">What this workspace delivers</div>
                    <div class="hero-panel-grid">
                        <div class="hero-stat">
                            <div class="hero-stat-label">Health</div>
                            <div class="hero-stat-value">Data + retail quality</div>
                        </div>
                        <div class="hero-stat">
                            <div class="hero-stat-label">Signals</div>
                            <div class="hero-stat-value">Trend + momentum views</div>
                        </div>
                        <div class="hero-stat">
                            <div class="hero-stat-label">Gaps</div>
                            <div class="hero-stat-value">Distribution whitespace</div>
                        </div>
                        <div class="hero-stat">
                            <div class="hero-stat-label">Action</div>
                            <div class="hero-stat-value">Workbook + PDF export</div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.markdown(
        f"<div class='small-note' style='margin:0.15rem 0 0.7rem 0;'>{APP_SUBTITLE}</div>",
        unsafe_allow_html=True
    )

def metric_card(label: str, value: str, sub: str = ""):
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{label}</div>
            <div class="metric-value">{value}</div>
            <div class="metric-sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# =========================================================
# DATA HELPERS
# =========================================================
def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [
        str(c).strip().lower().replace(" ", "_").replace("-", "_")
        for c in out.columns
    ]
    return out

def normalize_state(x):
    if pd.isna(x):
        return np.nan
    return str(x).strip().upper()

def read_excel_sheet(uploaded_file, sheet_name: str) -> pd.DataFrame:
    uploaded_file.seek(0)
    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=0)
    return normalize_columns(df)

def read_uploaded_table(uploaded_file):
    if uploaded_file is None:
        return None
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return normalize_columns(pd.read_csv(uploaded_file))
    if name.endswith(".xlsx") or name.endswith(".xls"):
        uploaded_file.seek(0)
        return normalize_columns(pd.read_excel(uploaded_file, header=0))
    raise ValueError(f"Unsupported file type: {uploaded_file.name}")

def validate_required_columns(products, stores, sales_history, shelf=None):
    required_product_cols = {"sku_id"}
    required_store_cols = {"store_id", "retailer", "state"}
    required_sales_cols = {"store_id", "sku_id", "week_end_date", "units"}

    missing = {}
    if not required_product_cols.issubset(products.columns):
        missing["Products"] = sorted(list(required_product_cols - set(products.columns)))
    if not required_store_cols.issubset(stores.columns):
        missing["Stores"] = sorted(list(required_store_cols - set(stores.columns)))
    if not required_sales_cols.issubset(sales_history.columns):
        missing["Sales_History"] = sorted(list(required_sales_cols - set(sales_history.columns)))
    if shelf is not None and len(shelf) > 0:
        required_shelf_cols = {"store_id", "sku_id"}
        if not required_shelf_cols.issubset(shelf.columns):
            missing["Shelf_Snapshot"] = sorted(list(required_shelf_cols - set(shelf.columns)))
    return missing

def classify_data_quality_score(score: float) -> str:
    if score >= 95:
        return "Excellent"
    if score >= 85:
        return "Good"
    if score >= 70:
        return "Fair"
    return "Needs Cleanup"

def classify_health_score(score: float) -> str:
    if score >= 90:
        return "Excellent"
    if score >= 75:
        return "Strong"
    if score >= 60:
        return "Fair"
    return "Weak"

def prepare_stores(stores: pd.DataFrame) -> pd.DataFrame:
    stores = normalize_columns(stores).copy()
    stores["store_id"] = stores["store_id"].astype(str).str.strip()
    stores["retailer"] = stores["retailer"].astype(str).str.strip()
    if "format" not in stores.columns:
        stores["format"] = "Unknown"
    else:
        stores["format"] = stores["format"].fillna("Unknown").astype(str).str.strip()
    if "store_name" not in stores.columns:
        stores["store_name"] = np.nan
    stores["state"] = stores["state"].apply(normalize_state)
    stores["region"] = stores["state"].map(STATE_TO_REGION)
    return stores

# =========================================================
# DATA QUALITY ENGINE
# =========================================================
def run_data_quality_checks(products, stores, sales_history, shelf=None):
    issues = []

    products = normalize_columns(products).copy()
    stores = normalize_columns(stores).copy()
    sales_history = normalize_columns(sales_history).copy()

    if shelf is not None and len(shelf) > 0:
        shelf = normalize_columns(shelf).copy()
    else:
        shelf = pd.DataFrame()

    rows_uploaded = len(sales_history)

    if "store_id" in stores.columns:
        stores["store_id"] = stores["store_id"].astype(str).str.strip()
    if "sku_id" in products.columns:
        products["sku_id"] = products["sku_id"].astype(str).str.strip()
    if "store_id" in sales_history.columns:
        sales_history["store_id"] = sales_history["store_id"].astype(str).str.strip()
    if "sku_id" in sales_history.columns:
        sales_history["sku_id"] = sales_history["sku_id"].astype(str).str.strip()
    if "state" in stores.columns:
        stores["state"] = stores["state"].apply(normalize_state)

    def add_issue(check, status, count, severity):
        issues.append({
            "check": check,
            "status": status,
            "count": int(count),
            "severity_weight": severity
        })

    missing_store = sales_history["store_id"].isna().sum() + (sales_history["store_id"].astype(str).str.strip() == "").sum()
    add_issue("Missing store_id in Sales_History", "Fail" if missing_store > 0 else "Pass", missing_store, 8)

    missing_sku = sales_history["sku_id"].isna().sum() + (sales_history["sku_id"].astype(str).str.strip() == "").sum()
    add_issue("Missing sku_id in Sales_History", "Fail" if missing_sku > 0 else "Pass", missing_sku, 8)

    parsed_dates = pd.to_datetime(sales_history["week_end_date"], errors="coerce")
    invalid_dates = parsed_dates.isna().sum()
    sales_history["week_end_date"] = parsed_dates
    add_issue("Invalid week_end_date values", "Fail" if invalid_dates > 0 else "Pass", invalid_dates, 7)

    sales_history["units"] = pd.to_numeric(sales_history["units"], errors="coerce")
    non_numeric_units = sales_history["units"].isna().sum()
    add_issue("Non-numeric units", "Fail" if non_numeric_units > 0 else "Pass", non_numeric_units, 7)
    sales_history["units"] = sales_history["units"].fillna(0)

    negative_units = (sales_history["units"] < 0).sum()
    negative_units_pct = round((negative_units / max(len(sales_history), 1)) * 100, 2)
    neg_units_status = "Pass"
    if negative_units_pct > 20:
        neg_units_status = "Fail"
    elif negative_units_pct > 5:
        neg_units_status = "Warn"
    add_issue("Negative units present", neg_units_status, negative_units, 4)

    sales_col = None
    if "sales_dollars" in sales_history.columns:
        sales_col = "sales_dollars"
    elif "sales" in sales_history.columns:
        sales_col = "sales"

    negative_sales_pct = 0.0
    if sales_col is not None:
        sales_history[sales_col] = pd.to_numeric(sales_history[sales_col], errors="coerce")
        non_numeric_sales = sales_history[sales_col].isna().sum()
        add_issue(f"Non-numeric {sales_col}", "Fail" if non_numeric_sales > 0 else "Pass", non_numeric_sales, 7)
        sales_history[sales_col] = sales_history[sales_col].fillna(0)

        negative_sales = (sales_history[sales_col] < 0).sum()
        negative_sales_pct = round((negative_sales / max(len(sales_history), 1)) * 100, 2)
        neg_sales_status = "Pass"
        if negative_sales_pct > 20:
            neg_sales_status = "Fail"
        elif negative_sales_pct > 5:
            neg_sales_status = "Warn"
        add_issue(f"Negative {sales_col} present", neg_sales_status, negative_sales, 4)

    dup_count = sales_history.duplicated(subset=["store_id", "sku_id", "week_end_date"]).sum()
    add_issue("Duplicate store_id + sku_id + week_end_date rows", "Warn" if dup_count > 0 else "Pass", dup_count, 3)

    unmatched_skus = (~sales_history["sku_id"].isin(products["sku_id"].astype(str).str.strip())).sum()
    add_issue("Sales_History sku_id not found in Products", "Fail" if unmatched_skus > 0 else "Pass", unmatched_skus, 6)

    unmatched_stores = (~sales_history["store_id"].isin(stores["store_id"].astype(str).str.strip())).sum()
    add_issue("Sales_History store_id not found in Stores", "Fail" if unmatched_stores > 0 else "Pass", unmatched_stores, 6)

    invalid_states = (~stores["state"].isin(list(STATE_TO_REGION.keys()))).sum()
    add_issue("Invalid state codes in Stores", "Fail" if invalid_states > 0 else "Pass", invalid_states, 5)

    unmapped_region_count = stores["state"].map(STATE_TO_REGION).isna().sum()
    add_issue("States that could not be mapped to region", "Fail" if unmapped_region_count > 0 else "Pass", unmapped_region_count, 5)

    sparse_pairs = 0
    if {"store_id", "sku_id", "week_end_date"}.issubset(sales_history.columns):
        counts = sales_history.groupby(["store_id", "sku_id"])["week_end_date"].nunique().reset_index(name="week_count")
        sparse_pairs = int((counts["week_count"] < counts["week_count"].median()).sum()) if len(counts) else 0
        add_issue("Store/SKU pairs with below-median week coverage", "Warn" if sparse_pairs > 0 else "Pass", sparse_pairs, 2)

    if len(shelf) > 0:
        if "shelf_share" in shelf.columns:
            shelf["shelf_share"] = pd.to_numeric(shelf["shelf_share"], errors="coerce")
            bad_shelf_share = ((shelf["shelf_share"] < 0) | (shelf["shelf_share"] > 1)).sum()
            add_issue("Shelf_Snapshot shelf_share outside 0 to 1", "Fail" if bad_shelf_share > 0 else "Pass", bad_shelf_share, 5)
        if "facings" in shelf.columns:
            shelf["facings"] = pd.to_numeric(shelf["facings"], errors="coerce")
            bad_facings = (shelf["facings"] < 0).sum()
            add_issue("Negative facings in Shelf_Snapshot", "Fail" if bad_facings > 0 else "Pass", bad_facings, 5)

    quality = pd.DataFrame(issues)

    penalty = 0
    for _, row in quality.iterrows():
        if row["status"] == "Fail":
            penalty += min(row["count"], 10) * row["severity_weight"]
        elif row["status"] == "Warn":
            penalty += min(row["count"], 10) * row["severity_weight"] * 0.35

    data_quality_score = max(0, round(100 - penalty, 1))

    rejected_mask = (
        sales_history["store_id"].isna() |
        (sales_history["store_id"].astype(str).str.strip() == "") |
        sales_history["sku_id"].isna() |
        (sales_history["sku_id"].astype(str).str.strip() == "") |
        sales_history["week_end_date"].isna()
    )

    rows_rejected = int(rejected_mask.sum())
    rows_accepted = int(rows_uploaded - rows_rejected)

    abs_total_units = float(sales_history["units"].abs().sum())
    abs_negative_units = float(sales_history.loc[sales_history["units"] < 0, "units"].abs().sum())
    return_impact_score = round((abs_negative_units / abs_total_units) * 100, 2) if abs_total_units > 0 else 0.0

    meta = {
        "rows_uploaded": rows_uploaded,
        "rows_accepted": rows_accepted,
        "rows_rejected": rows_rejected,
        "negative_units_pct": negative_units_pct,
        "negative_sales_pct": negative_sales_pct,
        "return_impact_score": return_impact_score,
        "data_quality_score": data_quality_score,
        "quality_label": classify_data_quality_score(data_quality_score),
    }
    return quality, meta

# =========================================================
# AI INSIGHTS ENGINE
# =========================================================
def build_ai_insights(summary, underperf, dist, yoy, momentum, shelf_df, quality):
    insights = []

    if len(underperf):
        row = underperf.sort_values("revenue_opportunity_score", ascending=False).iloc[0]
        insights.append(
            f"Biggest store execution issue is store {row['store_id']} at {row['retailer']} in {row['region']}, "
            f"with SPI {row['store_performance_index']:.1f} and opportunity ${row['revenue_opportunity_score']:,.0f}."
        )

    if len(dist):
        row = dist.sort_values("distribution_gap_count", ascending=False).iloc[0]
        insights.append(
            f"Top distribution expansion opportunity is {row['brand']} in {row['retailer']}, "
            f"with a gap of {int(row['distribution_gap_count'])} stores."
        )

    if len(yoy):
        yoy_clean = yoy.dropna(subset=["yoy_sales_growth_pct"])
        if len(yoy_clean):
            top = yoy_clean.sort_values("yoy_sales_growth_pct", ascending=False).iloc[0]
            bottom = yoy_clean.sort_values("yoy_sales_growth_pct", ascending=True).iloc[0]
            insights.append(
                f"Top YoY winner is {top['sku_id']} ({top['brand']}) at {top['yoy_sales_growth_pct']:.1f}% sales growth."
            )
            insights.append(
                f"Biggest YoY decliner is {bottom['sku_id']} ({bottom['brand']}) at {bottom['yoy_sales_growth_pct']:.1f}%."
            )

    if len(momentum):
        up = momentum[momentum["momentum_flag"] == "Trending Up"]
        down = momentum[momentum["momentum_flag"] == "Trending Down"]
        if len(up):
            row = up.sort_values("momentum_ratio", ascending=False).iloc[0]
            insights.append(
                f"Strongest momentum SKU is {row['sku_id']} ({row['brand']}) with ratio {row['momentum_ratio']:.2f}."
            )
        if len(down):
            row = down.sort_values("momentum_ratio", ascending=True).iloc[0]
            insights.append(
                f"Momentum risk is {row['sku_id']} ({row['brand']}) with ratio {row['momentum_ratio']:.2f}."
            )

    if len(shelf_df):
        winners = shelf_df[shelf_df["shelf_action"] == "Increase Facings"]
        if len(winners):
            row = winners.sort_values("space_efficiency_index", ascending=False).iloc[0]
            insights.append(
                f"Shelf space winner is {row['sku_id']} ({row['brand']}) with SEI {row['space_efficiency_index']:.1f}."
            )

    fail_count = int((quality["status"] == "Fail").sum()) if len(quality) else 0
    warn_count = int((quality["status"] == "Warn").sum()) if len(quality) else 0
    insights.append(
        f"Data quality produced {fail_count} failures and {warn_count} warnings; return impact is {summary['return_impact_score']}%."
    )

    return pd.DataFrame({"ai_insight": insights[:8]})

# =========================================================
# RECOMMENDATIONS ENGINE
# =========================================================
def build_recommendations(underperf, dist, yoy, momentum, shelf_df):
    recs = []

    if len(underperf):
        row = underperf.sort_values("revenue_opportunity_score", ascending=False).iloc[0]
        recs.append(
            f"Investigate store {row['store_id']} at {row['retailer']} in {row['region']}: "
            f"revenue opportunity is ${row['revenue_opportunity_score']:,.0f} and SPI is {row['store_performance_index']:.1f}."
        )

    if len(dist):
        row = dist.sort_values("distribution_gap_count", ascending=False).iloc[0]
        recs.append(
            f"Expand distribution for {row['brand']} / {row['category']} at {row['retailer']}: "
            f"gap of {int(row['distribution_gap_count'])} stores."
        )

    if len(yoy):
        yoy_clean = yoy.dropna(subset=["yoy_sales_growth_pct"])
        if len(yoy_clean):
            top = yoy_clean.sort_values("yoy_sales_growth_pct", ascending=False).iloc[0]
            bottom = yoy_clean.sort_values("yoy_sales_growth_pct", ascending=True).iloc[0]
            recs.append(f"Protect and expand SKU {top['sku_id']} ({top['brand']}): YoY sales growth is {top['yoy_sales_growth_pct']:.1f}%.")
            recs.append(f"Review SKU {bottom['sku_id']} ({bottom['brand']}): YoY sales growth is {bottom['yoy_sales_growth_pct']:.1f}%.")

    if len(momentum):
        up = momentum[momentum["momentum_flag"] == "Trending Up"]
        down = momentum[momentum["momentum_flag"] == "Trending Down"]
        if len(up):
            row = up.sort_values("momentum_ratio", ascending=False).iloc[0]
            recs.append(f"Increase support behind momentum winner {row['sku_id']} ({row['brand']}): momentum ratio is {row['momentum_ratio']:.2f}.")
        if len(down):
            row = down.sort_values("momentum_ratio", ascending=True).iloc[0]
            recs.append(f"Diagnose decline for {row['sku_id']} ({row['brand']}): momentum ratio is {row['momentum_ratio']:.2f}.")

    if len(shelf_df):
        winners = shelf_df[shelf_df["shelf_action"] == "Increase Facings"]
        losers = shelf_df[shelf_df["shelf_action"] == "Reduce / Review"]
        if len(winners):
            row = winners.sort_values("space_efficiency_index", ascending=False).iloc[0]
            recs.append(f"Increase facings for SKU {row['sku_id']} ({row['brand']}): Space Efficiency Index is {row['space_efficiency_index']:.1f}.")
        if len(losers):
            row = losers.sort_values("space_efficiency_index", ascending=True).iloc[0]
            recs.append(f"Review shelf space for SKU {row['sku_id']} ({row['brand']}): Space Efficiency Index is {row['space_efficiency_index']:.1f}.")

    return pd.DataFrame({"recommended_action": recs[:8]})

# =========================================================
# SELL-IN ENGINE
# =========================================================
def build_sell_in_engine(dist, momentum, yoy, shelf_df, underperf):
    rows = []

    if len(dist):
        dist_top = dist.sort_values(["distribution_gap_count", "distribution_gap_index"], ascending=[False, False]).head(10)
        for _, row in dist_top.iterrows():
            rows.append({
                "priority": "High" if row["distribution_gap_count"] >= 10 else "Medium",
                "retailer": row["retailer"],
                "sku_or_brand": row["brand"],
                "action": "Expand distribution",
                "rationale": f"Distribution gap of {int(row['distribution_gap_count'])} stores in {row['retailer']}.",
                "estimated_opportunity": np.nan
            })

    if len(shelf_df):
        winners = shelf_df[shelf_df["shelf_action"] == "Increase Facings"].sort_values("space_efficiency_index", ascending=False).head(10)
        for _, row in winners.iterrows():
            rows.append({
                "priority": "High" if row["space_efficiency_index"] >= 140 else "Medium",
                "retailer": row.get("retailer", "Mixed"),
                "sku_or_brand": row["sku_id"],
                "action": "Increase facings",
                "rationale": f"Space Efficiency Index of {row['space_efficiency_index']:.1f} with strong shelf productivity.",
                "estimated_opportunity": row.get("total_sales", np.nan)
            })

    if len(underperf):
        top_exec = underperf.sort_values("revenue_opportunity_score", ascending=False).head(10)
        for _, row in top_exec.iterrows():
            rows.append({
                "priority": row["opportunity_priority"],
                "retailer": row["retailer"],
                "sku_or_brand": row["store_id"],
                "action": "Fix store execution",
                "rationale": f"SPI of {row['store_performance_index']:.1f} and revenue opportunity of ${row['revenue_opportunity_score']:,.0f}.",
                "estimated_opportunity": row["revenue_opportunity_score"]
            })

    if len(momentum):
        movers = momentum[momentum["momentum_flag"] == "Trending Up"].sort_values("momentum_ratio", ascending=False).head(10)
        for _, row in movers.iterrows():
            rows.append({
                "priority": "Medium",
                "retailer": "Mixed",
                "sku_or_brand": row["sku_id"],
                "action": "Sell-in support",
                "rationale": f"Momentum ratio of {row['momentum_ratio']:.2f} indicates strong recent acceleration.",
                "estimated_opportunity": np.nan
            })

    sell_in = pd.DataFrame(rows)
    if len(sell_in):
        sell_in = sell_in.drop_duplicates(subset=["retailer", "sku_or_brand", "action"]).reset_index(drop=True)
    return sell_in

# =========================================================
# MAIN ANALYTICS ENGINE
# =========================================================
def run_analysis(products, stores, sales_history, shelf=None):
    products = normalize_columns(products).copy()
    sales_history = normalize_columns(sales_history).copy()

    products["sku_id"] = products["sku_id"].astype(str).str.strip()
    if "brand" not in products.columns:
        products["brand"] = "Unknown"
    else:
        products["brand"] = products["brand"].fillna("Unknown")

    if "category" not in products.columns:
        products["category"] = "Unknown"
    else:
        products["category"] = products["category"].fillna("Unknown")

    stores = prepare_stores(stores)

    sales_history["store_id"] = sales_history["store_id"].astype(str).str.strip()
    sales_history["sku_id"] = sales_history["sku_id"].astype(str).str.strip()
    sales_history["week_end_date"] = pd.to_datetime(sales_history["week_end_date"], errors="coerce")
    sales_history["units"] = pd.to_numeric(sales_history["units"], errors="coerce").fillna(0)

    if "sales_dollars" in sales_history.columns:
        sales_history["sales_dollars"] = pd.to_numeric(sales_history["sales_dollars"], errors="coerce").fillna(0)
    elif "sales" in sales_history.columns:
        sales_history["sales_dollars"] = pd.to_numeric(sales_history["sales"], errors="coerce").fillna(0)
    else:
        sales_history["sales_dollars"] = 0.0

    if shelf is None:
        shelf = pd.DataFrame(columns=["store_id", "sku_id", "facings", "shelf_share"])
    else:
        shelf = normalize_columns(shelf).copy()
        if "store_id" in shelf.columns:
            shelf["store_id"] = shelf["store_id"].astype(str).str.strip()
        if "sku_id" in shelf.columns:
            shelf["sku_id"] = shelf["sku_id"].astype(str).str.strip()
        if "facings" in shelf.columns:
            shelf["facings"] = pd.to_numeric(shelf["facings"], errors="coerce").fillna(0)
        if "shelf_share" in shelf.columns:
            shelf["shelf_share"] = pd.to_numeric(shelf["shelf_share"], errors="coerce").fillna(0)

    quality_checks, quality_meta = run_data_quality_checks(products, stores, sales_history, shelf)

    sales_enriched = (
        sales_history
        .merge(products, on="sku_id", how="left")
        .merge(stores, on="store_id", how="left")
    )

    for col in ["brand", "category", "retailer", "format", "state", "region"]:
        if col not in sales_enriched.columns:
            sales_enriched[col] = "Unknown"
        else:
            sales_enriched[col] = sales_enriched[col].fillna("Unknown")

    current_max_week = sales_enriched["week_end_date"].max()
    if pd.isna(current_max_week):
        raise ValueError("No valid dates found in Sales_History.")

    trailing_13w_start = current_max_week - pd.Timedelta(weeks=13)
    trailing_52w_start = current_max_week - pd.Timedelta(weeks=52)

    sales_13w = sales_enriched[sales_enriched["week_end_date"] > trailing_13w_start].copy()
    sales_52w = sales_enriched[sales_enriched["week_end_date"] > trailing_52w_start].copy()

    weeks_13 = max(sales_13w["week_end_date"].nunique(), 1)
    weeks_52 = max(sales_52w["week_end_date"].nunique(), 1)

    # SKU Velocity
    sku_velocity = (
        sales_13w.groupby(["sku_id", "brand", "category"], dropna=False)
        .agg(
            total_units=("units", "sum"),
            total_sales=("sales_dollars", "sum"),
            active_stores=("store_id", "nunique")
        )
        .reset_index()
    )
    sku_velocity["velocity_units_per_store_per_week"] = (
        sku_velocity["total_units"] / sku_velocity["active_stores"].clip(lower=1) / weeks_13
    )
    category_avg_velocity = (
        sku_velocity.groupby("category", dropna=False)["velocity_units_per_store_per_week"]
        .mean()
        .rename("category_avg_velocity")
        .reset_index()
    )
    sku_velocity = sku_velocity.merge(category_avg_velocity, on="category", how="left")
    sku_velocity["sku_velocity_index"] = (
        sku_velocity["velocity_units_per_store_per_week"] /
        sku_velocity["category_avg_velocity"].replace(0, np.nan)
    ) * 100
    sku_velocity["sku_velocity_index"] = sku_velocity["sku_velocity_index"].fillna(0)

    # Store Performance Index + Revenue Opportunity
    store_totals = (
        sales_13w.groupby(["store_id", "retailer", "region", "state", "format"], dropna=False)
        .agg(
            actual_sales=("sales_dollars", "sum"),
            actual_units=("units", "sum"),
            sku_count=("sku_id", "nunique")
        )
        .reset_index()
    )

    peer_avg = (
        store_totals.groupby(["retailer", "format", "region"], dropna=False)["actual_sales"]
        .mean()
        .rename("expected_sales")
        .reset_index()
    )

    store_perf = store_totals.merge(peer_avg, on=["retailer", "format", "region"], how="left")
    store_perf["expected_sales"] = store_perf["expected_sales"].fillna(store_perf["actual_sales"].mean())
    store_perf["store_performance_index"] = (
        store_perf["actual_sales"] / store_perf["expected_sales"].replace(0, np.nan)
    ) * 100
    store_perf["store_performance_index"] = store_perf["store_performance_index"].fillna(0)
    store_perf["sales_gap"] = store_perf["expected_sales"] - store_perf["actual_sales"]
    store_perf["underperforming_flag"] = store_perf["store_performance_index"] < 80

    store_perf["opportunity_confidence"] = np.select(
        [
            (store_perf["store_performance_index"] < 70) & (quality_meta["data_quality_score"] >= 85),
            (store_perf["store_performance_index"] < 85) & (quality_meta["data_quality_score"] >= 70),
        ],
        ["High", "Medium"],
        default="Low"
    )

    confidence_factor = np.select(
        [
            store_perf["opportunity_confidence"] == "High",
            store_perf["opportunity_confidence"] == "Medium",
            store_perf["opportunity_confidence"] == "Low",
        ],
        [1.0, 0.7, 0.4],
        default=0.4
    )

    store_perf["revenue_opportunity_score"] = np.where(
        (store_perf["sales_gap"] > 0) & (store_perf["sales_gap"] >= 50),
        store_perf["sales_gap"] * confidence_factor,
        0
    )

    store_perf["opportunity_priority"] = np.select(
        [
            store_perf["revenue_opportunity_score"] >= 500,
            (store_perf["revenue_opportunity_score"] >= 200) & (store_perf["revenue_opportunity_score"] < 500),
            (store_perf["revenue_opportunity_score"] > 0) & (store_perf["revenue_opportunity_score"] < 200)
        ],
        ["High", "Medium", "Low"],
        default="None"
    )

    exception_flags = []
    for _, row in store_perf.iterrows():
        flags = []
        if row["store_performance_index"] < 70:
            flags.append("Store Execution Problem")
        if row["revenue_opportunity_score"] >= 500:
            flags.append("High Revenue Opportunity")
        exception_flags.append(" | ".join(flags) if flags else "Normal")
    store_perf["exception_flags"] = exception_flags

    underperforming_stores = (
        store_perf[store_perf["underperforming_flag"]]
        .sort_values(["revenue_opportunity_score", "store_performance_index"], ascending=[False, True])
        .reset_index(drop=True)
    )

    # Distribution Gap
    carried = (
        sales_13w.groupby(["brand", "category", "retailer", "store_id"], dropna=False)
        .agg(total_units=("units", "sum"))
        .reset_index()
    )
    carried = carried[carried["total_units"] > 0]

    retailer_store_universe = (
        stores.groupby("retailer", dropna=False)["store_id"]
        .nunique()
        .rename("retailer_store_universe")
        .reset_index()
    )

    distribution_gap = (
        carried.groupby(["brand", "category", "retailer"], dropna=False)["store_id"]
        .nunique()
        .rename("current_store_count")
        .reset_index()
        .merge(retailer_store_universe, on="retailer", how="left")
    )

    distribution_gap["distribution_gap_count"] = (
        distribution_gap["retailer_store_universe"] - distribution_gap["current_store_count"]
    ).clip(lower=0)
    distribution_gap["distribution_gap_index"] = (
        distribution_gap["distribution_gap_count"] /
        distribution_gap["retailer_store_universe"].replace(0, np.nan)
    ) * 100
    distribution_gap["distribution_gap_index"] = distribution_gap["distribution_gap_index"].fillna(0)

    # YoY
    sales_enriched["year"] = sales_enriched["week_end_date"].dt.year
    yearly_sku = (
        sales_enriched.groupby(["sku_id", "brand", "category", "year"], dropna=False)
        .agg(
            yearly_sales=("sales_dollars", "sum"),
            yearly_units=("units", "sum")
        )
        .reset_index()
    )

    years_available = sorted([y for y in yearly_sku["year"].dropna().unique()])
    yoy_growth = pd.DataFrame()

    if len(years_available) >= 2:
        prior_year = years_available[-2]
        current_year = years_available[-1]

        prior_df = yearly_sku[yearly_sku["year"] == prior_year][["sku_id", "yearly_sales", "yearly_units"]].rename(
            columns={"yearly_sales": f"sales_{prior_year}", "yearly_units": f"units_{prior_year}"}
        )
        current_df = yearly_sku[yearly_sku["year"] == current_year][["sku_id", "brand", "category", "yearly_sales", "yearly_units"]].rename(
            columns={"yearly_sales": f"sales_{current_year}", "yearly_units": f"units_{current_year}"}
        )

        yoy_growth = current_df.merge(prior_df, on="sku_id", how="left")
        yoy_growth[f"sales_{prior_year}"] = yoy_growth[f"sales_{prior_year}"].fillna(0)
        yoy_growth[f"units_{prior_year}"] = yoy_growth[f"units_{prior_year}"].fillna(0)

        yoy_growth["yoy_sales_growth_pct"] = np.where(
            yoy_growth[f"sales_{prior_year}"] > 0,
            ((yoy_growth[f"sales_{current_year}"] - yoy_growth[f"sales_{prior_year}"]) / yoy_growth[f"sales_{prior_year}"]) * 100,
            np.nan
        )
        yoy_growth["yoy_units_growth_pct"] = np.where(
            yoy_growth[f"units_{prior_year}"] > 0,
            ((yoy_growth[f"units_{current_year}"] - yoy_growth[f"units_{prior_year}"]) / yoy_growth[f"units_{prior_year}"]) * 100,
            np.nan
        )

        yoy_growth["exception_flags"] = np.select(
            [
                yoy_growth["yoy_sales_growth_pct"] >= 20,
                yoy_growth["yoy_sales_growth_pct"] <= -10
            ],
            [
                "YoY Winner",
                "YoY Decliner"
            ],
            default="Normal"
        )

    # Momentum
    velocity_13w = (
        sales_13w.groupby("sku_id", dropna=False)
        .agg(units_13w=("units", "sum"), stores_13w=("store_id", "nunique"))
        .reset_index()
    )
    velocity_13w["velocity_13w"] = velocity_13w["units_13w"] / velocity_13w["stores_13w"].clip(lower=1) / weeks_13

    velocity_52w = (
        sales_52w.groupby("sku_id", dropna=False)
        .agg(units_52w=("units", "sum"), stores_52w=("store_id", "nunique"))
        .reset_index()
    )
    velocity_52w["velocity_52w"] = velocity_52w["units_52w"] / velocity_52w["stores_52w"].clip(lower=1) / weeks_52

    momentum = (
        velocity_13w.merge(velocity_52w, on="sku_id", how="outer")
        .merge(products[["sku_id", "brand", "category"]], on="sku_id", how="left")
    )
    momentum["velocity_13w"] = momentum["velocity_13w"].fillna(0)
    momentum["velocity_52w"] = momentum["velocity_52w"].fillna(0)
    momentum["momentum_ratio"] = np.where(
        momentum["velocity_52w"] > 0,
        momentum["velocity_13w"] / momentum["velocity_52w"],
        np.nan
    )
    momentum["momentum_flag"] = np.select(
        [
            momentum["momentum_ratio"] >= 1.20,
            momentum["momentum_ratio"] <= 0.80
        ],
        ["Trending Up", "Trending Down"],
        default="Stable"
    )

    # Recent declines
    recent_declines = (
        sales_enriched.groupby(["sku_id", "brand", "category", "week_end_date"], dropna=False)
        .agg(weekly_sales=("sales_dollars", "sum"))
        .reset_index()
        .sort_values(["sku_id", "week_end_date"])
    )
    recent_declines["prev_week_sales"] = recent_declines.groupby("sku_id")["weekly_sales"].shift(1)
    recent_declines["wow_change_pct"] = (
        (recent_declines["weekly_sales"] - recent_declines["prev_week_sales"]) /
        recent_declines["prev_week_sales"].replace(0, np.nan)
    ) * 100
    recent_declines["wow_change_pct"] = recent_declines["wow_change_pct"].fillna(0)
    recent_declines = recent_declines[recent_declines["wow_change_pct"] <= -10].sort_values("wow_change_pct").reset_index(drop=True)

    # Shelf Productivity + SEI
    shelf_metrics = pd.DataFrame()
    if len(shelf) > 0 and {"store_id", "sku_id"}.issubset(shelf.columns):
        if "facings" not in shelf.columns:
            shelf["facings"] = np.nan
        if "shelf_share" not in shelf.columns:
            shelf["shelf_share"] = np.nan

        shelf_metrics = (
            shelf
            .merge(products, on="sku_id", how="left")
            .merge(
                sales_13w.groupby(["store_id", "sku_id"], dropna=False)
                .agg(total_sales=("sales_dollars", "sum"), total_units=("units", "sum"))
                .reset_index(),
                on=["store_id", "sku_id"],
                how="left"
            )
            .merge(stores[["store_id", "retailer", "state", "region", "format"]], on="store_id", how="left")
        )

        shelf_metrics["total_sales"] = shelf_metrics["total_sales"].fillna(0)
        shelf_metrics["total_units"] = shelf_metrics["total_units"].fillna(0)
        shelf_metrics["facings"] = pd.to_numeric(shelf_metrics["facings"], errors="coerce")
        shelf_metrics["shelf_productivity_score"] = (
            shelf_metrics["total_sales"] / shelf_metrics["facings"].replace(0, np.nan)
        )
        shelf_metrics["shelf_productivity_score"] = shelf_metrics["shelf_productivity_score"].fillna(0)
        shelf_metrics["sales_per_facing"] = shelf_metrics["shelf_productivity_score"]

        category_avg_spf = (
            shelf_metrics.groupby("category", dropna=False)["sales_per_facing"]
            .mean()
            .rename("category_avg_sales_per_facing")
            .reset_index()
        )
        shelf_metrics = shelf_metrics.merge(category_avg_spf, on="category", how="left")

        shelf_metrics["space_efficiency_index"] = (
            shelf_metrics["sales_per_facing"] /
            shelf_metrics["category_avg_sales_per_facing"].replace(0, np.nan)
        ) * 100
        shelf_metrics["space_efficiency_index"] = shelf_metrics["space_efficiency_index"].fillna(0)

        shelf_metrics["shelf_action"] = np.select(
            [
                shelf_metrics["space_efficiency_index"] >= 120,
                shelf_metrics["space_efficiency_index"] < 80
            ],
            ["Increase Facings", "Reduce / Review"],
            default="Hold"
        )

        shelf_metrics["exception_flags"] = np.select(
            [
                shelf_metrics["space_efficiency_index"] >= 120,
                shelf_metrics["space_efficiency_index"] < 80
            ],
            ["Shelf Space Winner", "Shelf Space Risk"],
            default="Normal"
        )

    # Recommendations / Sell-In / AI Insights
    recommendations = build_recommendations(underperforming_stores, distribution_gap, yoy_growth, momentum, shelf_metrics)
    sell_in = build_sell_in_engine(distribution_gap, momentum, yoy_growth, shelf_metrics, underperforming_stores)

    underperf_rate = float(store_perf["underperforming_flag"].mean()) if len(store_perf) else 0
    avg_spi = float(store_perf["store_performance_index"].fillna(100).mean()) if len(store_perf) else 100
    dist_gap_rate = float(distribution_gap["distribution_gap_index"].fillna(0).mean()) if len(distribution_gap) else 0

    a = min(max(avg_spi, 0), 120) / 120 * 40
    b = (1 - min(max(underperf_rate, 0), 1)) * 15
    c = (1 - min(max(dist_gap_rate / 100, 0), 1)) * 15
    d = (quality_meta["data_quality_score"] / 100) * 20
    e = min(quality_meta["return_impact_score"], 10) / 10 * 10
    retail_health_score = round(a + b + c + d + e, 1)

    health_summary = pd.DataFrame([{
        "retail_health_score": retail_health_score,
        "retail_health_label": classify_health_score(retail_health_score),
        "data_quality_score": quality_meta["data_quality_score"],
        "data_quality_label": quality_meta["quality_label"],
        "rows_uploaded": quality_meta["rows_uploaded"],
        "rows_accepted": quality_meta["rows_accepted"],
        "rows_rejected": quality_meta["rows_rejected"],
        "return_impact_score": quality_meta["return_impact_score"],
        "negative_units_pct": quality_meta["negative_units_pct"],
        "negative_sales_pct": quality_meta["negative_sales_pct"],
        "store_count": int(store_perf["store_id"].nunique()),
        "sku_count": int(products["sku_id"].nunique()),
        "underperforming_store_count": int(store_perf["underperforming_flag"].sum()),
        "avg_store_performance_index": round(avg_spi, 2),
        "avg_distribution_gap_index": round(dist_gap_rate, 2),
        "estimated_revenue_opportunity": round(float(store_perf["revenue_opportunity_score"].sum()), 2),
    }])

    ai_insights = build_ai_insights(
        health_summary.iloc[0],
        underperforming_stores,
        distribution_gap,
        yoy_growth,
        momentum,
        shelf_metrics,
        quality_checks
    )

    return {
        "health_summary": health_summary,
        "quality_checks": quality_checks,
        "ai_insights": ai_insights,
        "recommendations": recommendations,
        "sell_in_opportunities": sell_in,
        "store_performance_index": store_perf,
        "underperforming_stores": underperforming_stores,
        "sku_velocity_score": sku_velocity,
        "distribution_gap_index": distribution_gap,
        "recent_sku_declines": recent_declines,
        "shelf_productivity_score": shelf_metrics,
        "yoy_growth": yoy_growth,
        "momentum": momentum,
    }

# =========================================================
# EXPORT HELPERS
# =========================================================
def style_workbook_sheet(ws, title=None):
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if title:
        ws.insert_rows(1, 3)
        ws["B2"] = title
        ws["B2"].font = Font(size=16, bold=True)
        ws["B3"] = APP_SUBTITLE
        ws["B3"].font = Font(size=10, italic=True)
        add_logo_to_sheet(ws, st.session_state.get("logo_bytes"), "A1", width=95)
        header_row = 4
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 36)

def to_excel_download(results_dict, logo_bytes=None):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in results_dict.items():
            if isinstance(df, pd.DataFrame) and len(df) > 0:
                df.to_excel(writer, sheet_name=name[:31], index=False)
        workbook = writer.book
        for ws in workbook.worksheets:
            add_logo_to_sheet(ws, logo_bytes, "A1", width=95)
            ws.insert_rows(1, 3)
            ws["B2"] = f"{APP_TITLE} | {ws.title}"
            ws["B2"].font = Font(size=16, bold=True)
            ws["B3"] = APP_SUBTITLE
            ws["B3"].font = Font(size=10, italic=True)
            header_fill = PatternFill("solid", fgColor="0F172A")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col_cells in ws.columns:
                length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 36)
    output.seek(0)
    return output

def build_health_summary_workbook(health_df, quality_df, ai_df, recommendations_df, logo_bytes=None):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        health_df.to_excel(writer, sheet_name="Health_Summary", index=False)
        if quality_df is not None and len(quality_df):
            quality_df.to_excel(writer, sheet_name="Quality_Checks", index=False)
        if ai_df is not None and len(ai_df):
            ai_df.to_excel(writer, sheet_name="AI_Insights", index=False)
        if recommendations_df is not None and len(recommendations_df):
            recommendations_df.to_excel(writer, sheet_name="Recommendations", index=False)
        workbook = writer.book
        for ws in workbook.worksheets:
            add_logo_to_sheet(ws, logo_bytes, "A1", width=95)
            ws.insert_rows(1, 3)
            ws["B2"] = f"{APP_TITLE} Health Summary"
            ws["B2"].font = Font(size=16, bold=True)
            ws["B3"] = APP_SUBTITLE
            ws["B3"].font = Font(size=10, italic=True)
            header_fill = PatternFill("solid", fgColor="0F172A")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col_cells in ws.columns:
                length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 36)
    output.seek(0)
    return output

def build_executive_pdf(summary, recommendations_df, sell_in_df, logo_bytes=None):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    y = height - 50

    if logo_bytes:
        try:
            logo = ImageReader(BytesIO(logo_bytes))
            c.drawImage(logo, 40, y - 20, width=80, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass
    elif Path(LOGO_PATH).exists():
        try:
            logo = ImageReader(LOGO_PATH)
            c.drawImage(logo, 40, y - 20, width=80, preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    c.setFont("Helvetica-Bold", 20)
    c.drawString(140, y, "ShelfIQ 911 Executive Report")
    y -= 30
    c.setFont("Helvetica", 10)
    c.drawString(140, y, APP_SUBTITLE)

    y -= 35
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "Executive Summary")
    y -= 20
    c.setFont("Helvetica", 11)

    lines = [
        f"Retail Health Score: {summary['retail_health_score']} ({summary['retail_health_label']})",
        f"Data Quality Score: {summary['data_quality_score']} ({summary['data_quality_label']})",
        f"Underperforming Stores: {int(summary['underperforming_store_count'])}",
        f"Average SPI: {round(summary['avg_store_performance_index'], 1)}",
        f"Revenue Opportunity: ${summary['estimated_revenue_opportunity']:,.0f}",
        f"Rows Uploaded / Accepted / Rejected: {int(summary['rows_uploaded'])} / {int(summary['rows_accepted'])} / {int(summary['rows_rejected'])}",
        f"Return Impact Score: {summary['return_impact_score']}%",
    ]
    for line in lines:
        c.drawString(50, y, line)
        y -= 18

    y -= 10
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "Top Recommended Actions")
    y -= 18
    c.setFont("Helvetica", 10)
    if recommendations_df is not None and len(recommendations_df) > 0:
        for _, row in recommendations_df.head(6).iterrows():
            c.drawString(50, y, f"- {row['recommended_action']}"[:110])
            y -= 16
            if y < 80:
                c.showPage()
                y = height - 50
                if logo_bytes:
                    try:
                        logo = ImageReader(BytesIO(logo_bytes))
                        c.drawImage(logo, 40, y - 20, width=80, preserveAspectRatio=True, mask="auto")
                    except Exception:
                        pass
    else:
        c.drawString(50, y, "- No recommendations available.")
        y -= 16

    y -= 10
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "Top Sell-In Opportunities")
    y -= 18
    c.setFont("Helvetica", 10)
    if sell_in_df is not None and len(sell_in_df) > 0:
        for _, row in sell_in_df.head(6).iterrows():
            c.drawString(50, y, f"- {row['priority']} | {row['retailer']} | {row['sku_or_brand']} | {row['action']}"[:110])
            y -= 14
            c.drawString(60, y, f"Reason: {row['rationale']}"[:108])
            y -= 18
            if y < 80:
                c.showPage()
                y = height - 50
                if logo_bytes:
                    try:
                        logo = ImageReader(BytesIO(logo_bytes))
                        c.drawImage(logo, 40, y - 20, width=80, preserveAspectRatio=True, mask="auto")
                    except Exception:
                        pass
    else:
        c.drawString(50, y, "- No sell-in opportunities available.")

    c.save()
    buffer.seek(0)
    return buffer


# =========================================================
# DISPLAY / NARRATIVE HELPERS
# =========================================================
CHART_COLORS = {
    "navy": "#0b1f33",
    "blue": "#2563eb",
    "sky": "#38bdf8",
    "teal": "#0f766e",
    "green": "#16a34a",
    "amber": "#d97706",
    "rose": "#dc2626",
    "slate": "#475569",
}


def safe_summary_get(summary, primary_key, fallback_key=None, default=0):
    try:
        if isinstance(summary, dict):
            if primary_key in summary and pd.notna(summary.get(primary_key)):
                return summary.get(primary_key)
            if fallback_key and fallback_key in summary and pd.notna(summary.get(fallback_key)):
                return summary.get(fallback_key)
        elif hasattr(summary, "get"):
            val = summary.get(primary_key, None)
            if pd.notna(val):
                return val
            if fallback_key:
                val2 = summary.get(fallback_key, None)
                if pd.notna(val2):
                    return val2
        return default
    except Exception:
        return default

def format_metric_value(value, kind="number"):
    try:
        value = float(value)
    except Exception:
        return str(value)
    if kind == "currency":
        if abs(value) >= 1_000_000:
            return f"${value/1_000_000:.1f}M"
        if abs(value) >= 1_000:
            return f"${value/1_000:.1f}K"
        return f"${value:,.0f}"
    if kind == "pct":
        return f"{value:.1f}%"
    if abs(value) >= 1_000_000:
        return f"{value/1_000_000:.1f}M"
    if abs(value) >= 1_000:
        return f"{value/1_000:.1f}K"
    return f"{value:,.1f}"

def status_badge(text, tone="good"):
    tone = tone if tone in {"good", "warn", "risk"} else "good"
    return f"<span class='badge badge-{tone}'>{text}</span>"


def apply_pro_theme(fig, title):
    fig.update_layout(
        title=dict(text=title, x=0, xanchor="left", font=dict(size=19, color="#0f172a")),
        height=410,
        margin=dict(l=20, r=20, t=72, b=24),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(255,255,255,0.92)",
        font=dict(color="#475467", size=12),
        hoverlabel=dict(bgcolor="white", font_color="#0f172a"),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1,
            bgcolor="rgba(255,255,255,0.8)",
            bordercolor="#e5ebf3",
            borderwidth=1
        ),
        xaxis=dict(
            showgrid=False,
            zeroline=False,
            showline=True,
            linecolor="#d9e2ec",
            tickfont=dict(color="#475467")
        ),
        yaxis=dict(
            showgrid=True,
            gridcolor="#eef2f6",
            gridwidth=1,
            zeroline=False,
            showline=False,
            tickfont=dict(color="#475467")
        )
    )
    fig.update_annotations(font_color="#475467")
    return fig

def chart_panel(fig):
    st.markdown("<div class='panel'>", unsafe_allow_html=True)
    st.plotly_chart(fig, use_container_width=True)
    st.markdown("</div>", unsafe_allow_html=True)


def render_insight_card(title, headline, what_happened, why_it_matters, action):
    st.markdown(
        f"""
        <div class="insight-card">
            <div class="insight-header">
                <div>
                    <div class="insight-kicker">{title}</div>
                    <div class="insight-headline">{headline}</div>
                </div>
                <div class="insight-sub">Business interpretation</div>
            </div>
            <div class="insight-grid">
                <div class="insight-pill">
                    <div class="insight-pill-label">What happened</div>
                    <div class="insight-pill-value">{what_happened}</div>
                </div>
                <div class="insight-pill">
                    <div class="insight-pill-label">Why it matters</div>
                    <div class="insight-pill-value">{why_it_matters}</div>
                </div>
                <div class="insight-pill">
                    <div class="insight-pill-label">Recommended action</div>
                    <div class="insight-pill-value">{action}</div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

def render_kpi_strip(items):
    cols = st.columns(len(items))
    for col, item in zip(cols, items):
        with col:
            metric_card(item.get("label","Metric"), item.get("value","-"), item.get("sub",""))

def best_row(df, metric, ascending=False):
    if df is None or len(df) == 0 or metric not in df.columns:
        return None
    temp = df.dropna(subset=[metric]).sort_values(metric, ascending=ascending)
    if len(temp) == 0:
        return None
    return temp.iloc[0]


# =========================================================
# DASHBOARD CHART HELPERS
# =========================================================

def bar_chart(df, x, y, title, color=None, top_n=10, ascending=False, horizontal=True):
    if df is None or len(df) == 0 or x not in df.columns or y not in df.columns:
        return None
    temp = df.copy().dropna(subset=[x, y]).sort_values(y, ascending=ascending).head(top_n)
    if len(temp) == 0:
        return None

    palette = [CHART_COLORS["blue"], CHART_COLORS["teal"], CHART_COLORS["amber"], CHART_COLORS["rose"], CHART_COLORS["slate"]]

    if horizontal:
        temp = temp.sort_values(y, ascending=True)
        fig = px.bar(
            temp, x=y, y=x, color=color, orientation="h", template="plotly_white",
            color_discrete_sequence=palette
        )
        if color is None:
            fig.update_traces(marker_color=CHART_COLORS["blue"])
        fig.update_traces(
            texttemplate="%{x:,.1f}",
            textposition="outside",
            cliponaxis=False,
            marker_line_color="rgba(255,255,255,0.95)",
            marker_line_width=1.2,
            hovertemplate="<b>%{y}</b><br>%{x:,.2f}<extra></extra>"
        )
    else:
        fig = px.bar(
            temp, x=x, y=y, color=color, template="plotly_white",
            color_discrete_sequence=palette
        )
        if color is None:
            fig.update_traces(marker_color=CHART_COLORS["blue"])
        fig.update_traces(
            texttemplate="%{y:,.1f}",
            textposition="outside",
            cliponaxis=False,
            marker_line_color="rgba(255,255,255,0.95)",
            marker_line_width=1.2,
            hovertemplate="<b>%{x}</b><br>%{y:,.2f}<extra></extra>"
        )

    fig.update_layout(bargap=0.28)
    return apply_pro_theme(fig, title)


def line_chart(df, x, y, title, color=None):
    if df is None or len(df) == 0 or x not in df.columns or y not in df.columns:
        return None
    temp = df.copy().dropna(subset=[x, y]).sort_values(x)
    if len(temp) == 0:
        return None

    if color and color in temp.columns:
        fig = px.line(
            temp, x=x, y=y, color=color, template="plotly_white", markers=True,
            color_discrete_sequence=[CHART_COLORS["blue"], CHART_COLORS["teal"], CHART_COLORS["amber"], CHART_COLORS["rose"]]
        )
        fig.update_traces(line=dict(width=3), marker=dict(size=6))
    else:
        fig = go.Figure()
        fig.add_trace(
            go.Scatter(
                x=temp[x],
                y=temp[y],
                mode="lines+markers",
                line=dict(width=3, color=CHART_COLORS["blue"], shape="spline", smoothing=0.55),
                marker=dict(size=6, color="#ffffff", line=dict(width=2, color=CHART_COLORS["blue"])),
                fill="tozeroy",
                fillcolor="rgba(29,78,216,0.10)",
                hovertemplate="<b>%{x}</b><br>%{y:,.2f}<extra></extra>"
            )
        )
    fig.update_layout(hovermode="x unified")
    return apply_pro_theme(fig, title)


def donut_chart(df, names, values, title):
    if df is None or len(df) == 0 or names not in df.columns or values not in df.columns:
        return None
    temp = df.copy().dropna(subset=[names, values])
    if len(temp) == 0:
        return None
    fig = px.pie(
        temp, names=names, values=values, hole=0.72, template="plotly_white",
        color_discrete_sequence=[CHART_COLORS["green"], CHART_COLORS["amber"], CHART_COLORS["rose"], CHART_COLORS["blue"], CHART_COLORS["slate"]]
    )
    fig.update_traces(
        textinfo="percent",
        textposition="inside",
        marker=dict(line=dict(color="white", width=2)),
        hovertemplate="<b>%{label}</b><br>%{value:,.0f} | %{percent}<extra></extra>"
    )
    return apply_pro_theme(fig, title)


def scatter_chart(df, x, y, title, color=None, hover_name=None, size=None):
    if df is None or len(df) == 0 or x not in df.columns or y not in df.columns:
        return None
    temp = df.copy().dropna(subset=[x, y])
    if len(temp) == 0:
        return None
    fig = px.scatter(
        temp, x=x, y=y, color=color, hover_name=hover_name, size=size, template="plotly_white",
        color_discrete_sequence=[CHART_COLORS["blue"], CHART_COLORS["teal"], CHART_COLORS["amber"], CHART_COLORS["rose"], CHART_COLORS["slate"]],
        opacity=0.86
    )
    fig.update_traces(
        marker=dict(size=12, line=dict(width=1.4, color="white")),
        hovertemplate="<b>%{hovertext}</b><br>X: %{x:,.2f}<br>Y: %{y:,.2f}<extra></extra>" if hover_name else None
    )
    return apply_pro_theme(fig, title)


def heatmap_chart(df, x, y, z, title):
    if df is None or len(df) == 0 or x not in df.columns or y not in df.columns or z not in df.columns:
        return None
    temp = df.pivot_table(index=y, columns=x, values=z, aggfunc="mean").fillna(0)
    if temp.empty:
        return None
    fig = go.Figure(
        data=go.Heatmap(
            z=temp.values,
            x=list(temp.columns),
            y=list(temp.index),
            colorscale=[[0, "#f8fbff"], [0.25, "#dbeafe"], [0.55, "#93c5fd"], [0.8, "#3b82f6"], [1, "#0f172a"]],
            hoverongaps=False,
            text=np.round(temp.values, 1),
            texttemplate="%{text}",
            colorbar=dict(thickness=14, outlinewidth=0)
        )
    )
    return apply_pro_theme(fig, title)

def build_tab_insight(title, body):
    render_insight_card(
        title=title,
        headline=body,
        what_happened="Performance patterns were ranked and surfaced from the current filtered view.",
        why_it_matters="This identifies where revenue, distribution, or execution is most at risk.",
        action="Prioritize the largest gap first and use the table below for store or SKU drill-down."
    )

def safe_top_value(df, sort_col, ascending=False, label_cols=None):
    if df is None or len(df) == 0 or sort_col not in df.columns:
        return "No standout issue detected."
    temp = df.dropna(subset=[sort_col]).sort_values(sort_col, ascending=ascending)
    if len(temp) == 0:
        return "No standout issue detected."
    row = temp.iloc[0]
    if not label_cols:
        return str(row.get(sort_col, ""))
    parts = [str(row[c]) for c in label_cols if c in row.index and pd.notna(row[c])]
    return " | ".join(parts) if parts else str(row.get(sort_col, ""))

# =========================================================
# APP LAYOUT
# =========================================================
if "logo_bytes" not in st.session_state:
    st.session_state["logo_bytes"] = get_logo_bytes()

render_header(st.session_state.get("logo_bytes"))
st.caption("Upload your retail files, validate structure, generate executive insights, and export polished outputs.")

with st.sidebar:
    uploaded_logo = st.file_uploader("Upload logo for dashboard + downloads", type=["png", "jpg", "jpeg"])
    if uploaded_logo is not None:
        st.session_state["logo_bytes"] = get_logo_bytes(uploaded_logo)
        save_logo_bytes(st.session_state["logo_bytes"])
        st.success("Logo loaded successfully.")
    elif st.session_state.get("logo_bytes"):
        st.caption("Using saved logo for on-screen view and downloads.")
    st.markdown("### Upload Mode")
    upload_mode = st.radio(
        "Choose upload method",
        ["One Excel workbook", "Separate files"],
        label_visibility="collapsed"
    )
    st.markdown("### Required Structure")
    st.markdown("""
- `Sales_History`
- `Products`
- `Stores`
- `Shelf_Snapshot` *(optional)*
""")

products = None
stores = None
sales_history = None
shelf = None

if upload_mode == "One Excel workbook":
    workbook = st.file_uploader(
        "Upload one Excel workbook with tabs: Sales_History, Products, Stores, and optional Shelf_Snapshot",
        type=["xlsx", "xls"]
    )
    if workbook is not None:
        try:
            products = read_excel_sheet(workbook, "Products")
            stores = read_excel_sheet(workbook, "Stores")
            sales_history = read_excel_sheet(workbook, "Sales_History")
            try:
                shelf = read_excel_sheet(workbook, "Shelf_Snapshot")
            except Exception:
                shelf = None

            st.success("Workbook loaded successfully.")
            with st.expander("Preview detected columns"):
                st.write("Products:", list(products.columns))
                st.write("Stores:", list(stores.columns))
                st.write("Sales_History:", list(sales_history.columns))
                if shelf is not None:
                    st.write("Shelf_Snapshot:", list(shelf.columns))
        except Exception as e:
            st.error(f"Could not read workbook: {e}")
else:
    c1, c2 = st.columns(2)
    with c1:
        sales_file = st.file_uploader("Upload sales history file", type=["csv", "xlsx", "xls"])
        products_file = st.file_uploader("Upload products file", type=["csv", "xlsx", "xls"])
    with c2:
        stores_file = st.file_uploader("Upload stores file", type=["csv", "xlsx", "xls"])
        shelf_file = st.file_uploader("Upload shelf file (optional)", type=["csv", "xlsx", "xls"])

    if sales_file is not None:
        sales_history = read_uploaded_table(sales_file)
    if products_file is not None:
        products = read_uploaded_table(products_file)
    if stores_file is not None:
        stores = read_uploaded_table(stores_file)
    if shelf_file is not None:
        shelf = read_uploaded_table(shelf_file)

run_clicked = st.button("Run ShelfIQ 911 Analysis", type="primary", use_container_width=True)

if run_clicked:
    if products is None or stores is None or sales_history is None:
        st.error("Please provide Products, Stores, and Sales_History data.")
        st.stop()

    try:
        missing = validate_required_columns(
            normalize_columns(products),
            normalize_columns(stores),
            normalize_columns(sales_history),
            normalize_columns(shelf) if shelf is not None else None
        )
        if missing:
            st.error(f"Missing required columns: {missing}")
            st.stop()

        results = run_analysis(products, stores, sales_history, shelf)

        health = results["health_summary"]
        quality = results["quality_checks"]
        ai_insights = results["ai_insights"]
        recommendations = results["recommendations"]
        sell_in = results["sell_in_opportunities"]
        underperf = results["underperforming_stores"]
        sku = results["sku_velocity_score"]
        dist = results["distribution_gap_index"]
        declines = results["recent_sku_declines"]
        shelf_df = results["shelf_productivity_score"]
        yoy = results["yoy_growth"]
        momentum = results["momentum"]

        summary = health.iloc[0]
        pdf_file = build_executive_pdf(summary, recommendations, sell_in, st.session_state.get("logo_bytes"))
        full_results_file = to_excel_download(results, st.session_state.get("logo_bytes"))

        st.success("Analysis complete.")

        fail_count = int((quality["status"] == "Fail").sum()) if len(quality) and "status" in quality.columns else 0
        warn_count = int((quality["status"] == "Warn").sum()) if len(quality) and "status" in quality.columns else 0
        data_quality_score = safe_summary_get(summary, "data_quality_score", "data_quality_score", 0)
        quality_label = safe_summary_get(summary, "data_quality_label", "quality_label", "Quality")
        revenue_oppty = safe_summary_get(summary, "estimated_revenue_opportunity", "revenue_opportunity_score", 0)

        st.markdown(
            f"""
            <div class="executive-band">
                <div class="executive-kicker">Performance command center</div>
                <div class="executive-title">Retail performance command story</div>
                <div class="executive-copy">
                    This version sharpens the overall experience with stronger KPI hierarchy, cleaner comparisons, easier signal reading, and more professional interpretation across data quality, store performance, whitespace, momentum, and shelf productivity.
                </div>
                <div class="signal-strip">
                    <div class="signal-card">
                        <div class="signal-label">Retail health</div>
                        <div class="signal-value">{summary['retail_health_score']} | {summary['retail_health_label']}</div>
                    </div>
                    <div class="signal-card">
                        <div class="signal-label">Revenue opportunity</div>
                        <div class="signal-value">${revenue_oppty:,.0f}</div>
                    </div>
                    <div class="signal-card">
                        <div class="signal-label">Underperforming stores</div>
                        <div class="signal-value">{int(summary['underperforming_store_count'])}</div>
                    </div>
                    <div class="signal-card">
                        <div class="signal-label">Data quality risk</div>
                        <div class="signal-value">{fail_count} fails | {warn_count} warnings</div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

        # KPI ROW
        k1, k2, k3, k4, k5 = st.columns(5)
        with k1:
            metric_card("Retail Health", f"{summary['retail_health_score']}", summary["retail_health_label"])
        with k2:
            metric_card("Data Quality", f"{data_quality_score}", quality_label)
        with k3:
            metric_card("Revenue Opportunity", f"${revenue_oppty:,.0f}", "Incremental upside")
        with k4:
            metric_card("Avg SPI", f"{summary['avg_store_performance_index']}", f"{int(summary['underperforming_store_count'])} underperforming stores")
        with k5:
            metric_card("Accepted Rows", f"{int(summary['rows_accepted']):,}", f"Return impact {summary['return_impact_score']}%")

        st.markdown("### Executive Narrative")
        left, right = st.columns([1.4, 1])

        with left:
            st.markdown("<div class='ai-box'>", unsafe_allow_html=True)
            st.markdown("**Insights Summary**")
            if len(ai_insights):
                cleaned_points = []
                for _, row in ai_insights.head(6).iterrows():
                    msg = str(row.get('ai_insight', '')).strip()
                    if msg:
                        cleaned_points.append(msg)
                if cleaned_points:
                    st.markdown("\n".join([f"- {pt}" for pt in cleaned_points]))
                else:
                    st.markdown("No AI insights generated.")
            else:
                st.markdown("No AI insights generated.")
            st.markdown("</div>", unsafe_allow_html=True)

        with right:
            st.markdown("<div class='panel'>", unsafe_allow_html=True)
            st.markdown("<div class='section-title'>Operational Integrity Snapshot</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='small-note'>Rows uploaded <b>{int(summary['rows_uploaded']):,}</b> · accepted <b>{int(summary['rows_accepted']):,}</b> · rejected <b>{int(summary['rows_rejected']):,}</b></div>", unsafe_allow_html=True)
            st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)
            st.markdown(f"{status_badge(f'{fail_count} failures', 'risk')} &nbsp; {status_badge(f'{warn_count} warnings', 'warn')} &nbsp; {status_badge(f'{quality_label} quality', 'good' if fail_count == 0 else 'warn')}", unsafe_allow_html=True)
            st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)
            st.markdown(f"- Negative units exposure: **{summary['negative_units_pct']}%**")
            st.markdown(f"- Negative sales exposure: **{summary['negative_sales_pct']}%**")
            st.markdown(f"- Avg store performance index: **{summary['avg_store_performance_index']}**")
            revenue_oppty = summary.get("estimated_revenue_opportunity", summary.get("revenue_opportunity_score", 0))
            st.markdown(f"- Revenue opportunity: **${revenue_oppty:,.0f}**")
            st.markdown("</div>", unsafe_allow_html=True)

        # CHARTS
        st.markdown("### Executive Performance Lens")
        r1c1, r1c2 = st.columns(2)
        with r1c1:
            if len(underperf):
                temp = underperf.copy()
                temp["store_label"] = temp["store_id"].astype(str)
                fig = bar_chart(temp, "store_label", "revenue_opportunity_score", "Top Store Revenue Opportunities", top_n=10)
                if fig:
                    st.plotly_chart(fig, use_container_width=True)
        with r1c2:
            if len(dist):
                temp = dist.copy()
                temp["gap_label"] = temp["brand"].astype(str) + " | " + temp["retailer"].astype(str)
                fig = bar_chart(temp, "gap_label", "distribution_gap_count", "Top Distribution Gaps", top_n=10)
                if fig:
                    st.plotly_chart(fig, use_container_width=True)

        r2c1, r2c2 = st.columns(2)
        with r2c1:
            if len(yoy):
                temp = yoy.dropna(subset=["yoy_sales_growth_pct"]).copy()
                if len(temp):
                    temp["sku_label"] = temp["sku_id"].astype(str)
                    fig = bar_chart(temp, "sku_label", "yoy_sales_growth_pct", "Top YoY Growth", top_n=10)
                    if fig:
                        st.plotly_chart(fig, use_container_width=True)
        with r2c2:
            if len(momentum):
                temp = momentum.copy()
                temp["sku_label"] = temp["sku_id"].astype(str)
                fig = bar_chart(temp, "sku_label", "momentum_ratio", "Top Momentum Movers", top_n=10)
                if fig:
                    st.plotly_chart(fig, use_container_width=True)

        # TREND LINE
        st.markdown("### Trend Lens")
        sales_trend = pd.DataFrame()
        if products is not None and stores is not None and sales_history is not None:
            sales_base = normalize_columns(sales_history).copy()
            stores_base = prepare_stores(stores)
            products_base = normalize_columns(products).copy()
            sales_base["store_id"] = sales_base["store_id"].astype(str).str.strip()
            sales_base["sku_id"] = sales_base["sku_id"].astype(str).str.strip()
            sales_base["units"] = pd.to_numeric(sales_base["units"], errors="coerce").fillna(0)
            if "sales_dollars" in sales_base.columns:
                sales_base["sales_dollars"] = pd.to_numeric(sales_base["sales_dollars"], errors="coerce").fillna(0)
            else:
                sales_base["sales_dollars"] = 0
            sales_base["week_end_date"] = pd.to_datetime(sales_base["week_end_date"], errors="coerce")
            sales_trend = (
                sales_base
                .merge(products_base[[c for c in ["sku_id", "brand", "category"] if c in products_base.columns]], on="sku_id", how="left")
                .merge(stores_base[["store_id", "retailer", "region"]], on="store_id", how="left")
                .groupby(["week_end_date", "retailer"], dropna=False)
                .agg(total_sales=("sales_dollars", "sum"), total_units=("units", "sum"))
                .reset_index()
                .dropna(subset=["week_end_date"])
            )

        if len(sales_trend):
            fig = line_chart(sales_trend, "week_end_date", "total_sales", "Weekly Sales Trend by Retailer", color="retailer")
            if fig:
                st.plotly_chart(fig, use_container_width=True)

        # ACTION PANELS
        st.markdown("### Action Layer")
        a1, a2 = st.columns(2)
        with a1:
            st.markdown("<div class='panel'>", unsafe_allow_html=True)
            st.markdown("<div class='section-title'>Recommendations</div>", unsafe_allow_html=True)
            if len(recommendations):
                st.dataframe(recommendations, use_container_width=True, hide_index=True)
            else:
                st.info("No recommendations available.")
            st.markdown("</div>", unsafe_allow_html=True)
        with a2:
            st.markdown("<div class='panel'>", unsafe_allow_html=True)
            st.markdown("<div class='section-title'>Sell-In Opportunities</div>", unsafe_allow_html=True)
            if len(sell_in):
                st.dataframe(sell_in, use_container_width=True, hide_index=True)
            else:
                st.info("No sell-in opportunities available.")
            st.markdown("</div>", unsafe_allow_html=True)


        # DETAIL TABS
        st.markdown("<div class='small-note' style='margin:0.2rem 0 0.55rem 0;'>Detailed measure tabs are below.</div>", unsafe_allow_html=True)
        tabs = st.tabs([
            "Data Quality",
            "Store Performance",
            "SKU Velocity",
            "Distribution Gaps",
            "YoY Growth",
            "Momentum",
            "Recent Declines",
            "Shelf Productivity"
        ])

        with tabs[0]:
            st.markdown("<div class='section-title'>Data Quality</div><div class='small-note'>Validate structural readiness and identify the highest-impact exceptions before acting on downstream measures.</div>", unsafe_allow_html=True)
            pass_checks = int((quality["status"] == "Pass").sum()) if len(quality) else 0
            warn_checks = int((quality["status"] == "Warn").sum()) if len(quality) else 0
            fail_checks = int((quality["status"] == "Fail").sum()) if len(quality) else 0
            render_kpi_strip([
                {"label": "Quality Score", "value": format_metric_value(summary["data_quality_score"]), "sub": f"{summary['data_quality_label']} data readiness"},
                {"label": "Failed Checks", "value": str(fail_checks), "sub": "Structural blockers to resolve"},
                {"label": "Warnings", "value": str(warn_checks), "sub": "Items that may distort analysis"},
            ])
            c1, c2 = st.columns([1, 1.25])
            with c1:
                if len(quality):
                    status_mix = quality.groupby("status", dropna=False).size().reset_index(name="count")
                    fig = donut_chart(status_mix, "status", "count", "Data Validation Mix")
                    if fig:
                        chart_panel(fig)
            with c2:
                fig = bar_chart(quality, "check", "count", "Most Material Data Exceptions", top_n=8)
                if fig:
                    chart_panel(fig)
            top_issue = best_row(quality, "count", ascending=False)
            headline = "Data quality is stable." if fail_checks == 0 else "Data quality needs cleanup before broad distribution decisions."
            what = f"Top issue: {top_issue['check']} ({int(top_issue['count'])})" if top_issue is not None else "No material issue identified."
            why = f"{fail_checks} failed checks and {warn_checks} warnings can suppress trust in downstream analytics."
            action = "Resolve failed checks first, then review warnings with the highest record counts."
            render_insight_card("Data Quality Readout", headline, what, why, action)
            st.dataframe(quality[["check", "status", "count"]], use_container_width=True, hide_index=True)

        with tabs[1]:
            st.markdown("<div class='section-title'>Store Performance</div><div class='small-note'>Compare expected versus actual store output and prioritize where the commercial gap is greatest.</div>", unsafe_allow_html=True)
            top_store = best_row(underperf, "revenue_opportunity_score", ascending=False)
            render_kpi_strip([
                {"label": "Largest Revenue Gap", "value": format_metric_value(top_store["revenue_opportunity_score"], "currency") if top_store is not None else "-", "sub": f"Store {top_store['store_id']}" if top_store is not None else "No store gap"},
                {"label": "Lowest SPI", "value": format_metric_value(underperf['store_performance_index'].min()) if len(underperf) and 'store_performance_index' in underperf.columns else "-", "sub": "Lowest indexed performer"},
                {"label": "Stores Flagged", "value": str(len(underperf)), "sub": "Locations needing action"},
            ])
            c1, c2 = st.columns(2)
            with c1:
                temp = underperf.copy()
                if len(temp):
                    temp["store_label"] = temp["store_id"].astype(str)
                    fig = scatter_chart(temp, "expected_sales", "actual_sales", "Expected vs. Actual Store Sales", color="retailer", hover_name="store_label", size="revenue_opportunity_score")
                    if fig:
                        fig.add_shape(type="line", x0=temp["expected_sales"].min(), y0=temp["expected_sales"].min(),
                                      x1=temp["expected_sales"].max(), y1=temp["expected_sales"].max(),
                                      line=dict(color="#94a3b8", dash="dash"))
                        chart_panel(fig)
            with c2:
                if len(underperf):
                    temp = underperf.copy()
                    temp["store_label"] = temp["store_id"].astype(str)
                    fig = bar_chart(temp, "store_label", "revenue_opportunity_score", "Highest Revenue Opportunity Stores", top_n=10)
                    if fig:
                        chart_panel(fig)
            heat = None
            if len(results["store_performance_index"]):
                heat = heatmap_chart(results["store_performance_index"], "retailer", "region", "store_performance_index", "Store Performance Index by Retailer and Region")
            if heat:
                chart_panel(heat)
            headline = "Underperformance is concentrated rather than broad-based."
            what = f"Primary opportunity sits in {safe_top_value(underperf, 'revenue_opportunity_score', ascending=False, label_cols=['retailer', 'store_id'])}."
            why = "Stores falling below expected sales represent the cleanest short-term revenue recovery opportunity."
            action = "Focus field execution, on-shelf availability, and local assortment in the top opportunity stores first."
            render_insight_card("Store Performance", headline, what, why, action)
            cols = [c for c in [
                "store_id", "retailer", "region", "actual_sales", "expected_sales",
                "store_performance_index", "sales_gap", "revenue_opportunity_score",
                "opportunity_priority", "opportunity_confidence", "exception_flags"
            ] if c in underperf.columns]
            st.dataframe(underperf[cols], use_container_width=True, hide_index=True)

        with tabs[2]:
            st.markdown("<div class='section-title'>SKU Velocity</div><div class='small-note'>Surface the fastest and weakest movers to guide assortment, replenishment, and space decisions.</div>", unsafe_allow_html=True)
            top_vel = best_row(sku, "velocity_units_per_store_per_week", ascending=False)
            render_kpi_strip([
                {"label": "Top Velocity SKU", "value": format_metric_value(top_vel["velocity_units_per_store_per_week"]) if top_vel is not None else "-", "sub": f"{top_vel['brand']} | {top_vel['sku_id']}" if top_vel is not None else "No velocity leader"},
                {"label": "Avg Category Velocity", "value": format_metric_value(sku["category_avg_velocity"].mean()) if len(sku) and "category_avg_velocity" in sku.columns else "-", "sub": "Cross-category benchmark"},
                {"label": "SKU Count", "value": str(len(sku)), "sub": "Items in current cut"},
            ])
            c1, c2 = st.columns(2)
            with c1:
                temp = sku.sort_values("velocity_units_per_store_per_week", ascending=False).head(15).copy() if len(sku) else pd.DataFrame()
                if len(temp):
                    temp["sku_label"] = temp["sku_id"].astype(str)
                    fig = bar_chart(temp, "sku_label", "velocity_units_per_store_per_week", "Highest Velocity SKUs", top_n=15)
                    if fig:
                        chart_panel(fig)
            with c2:
                if len(sku):
                    brand_velocity = sku.groupby("brand", dropna=False)["velocity_units_per_store_per_week"].sum().reset_index().sort_values("velocity_units_per_store_per_week", ascending=False).head(10)
                    fig = bar_chart(brand_velocity, "brand", "velocity_units_per_store_per_week", "Brand Contribution to Velocity", top_n=10)
                    if fig:
                        chart_panel(fig)
            headline = "Velocity leaders are clear and should anchor assortment decisions."
            what = f"Fastest item is {safe_top_value(sku, 'velocity_units_per_store_per_week', ascending=False, label_cols=['brand', 'sku_id'])}."
            why = "High-velocity SKUs typically justify stronger shelf presence and broader distribution support."
            action = "Protect in-stock, evaluate facings uplift, and use velocity leaders as the benchmark for slower items."
            render_insight_card("SKU Velocity", headline, what, why, action)
            cols = [c for c in [
                "sku_id", "brand", "category", "total_units", "active_stores",
                "velocity_units_per_store_per_week", "category_avg_velocity", "sku_velocity_index"
            ] if c in sku.columns]
            st.dataframe(sku.sort_values("velocity_units_per_store_per_week", ascending=False)[cols], use_container_width=True, hide_index=True)

        with tabs[3]:
            st.markdown("<div class='section-title'>Distribution Gaps</div><div class='small-note'>Pinpoint white space by brand, retailer, and category to support placement and sell-in planning.</div>", unsafe_allow_html=True)
            top_gap = best_row(dist, "distribution_gap_count", ascending=False)
            render_kpi_strip([
                {"label": "Largest Gap", "value": format_metric_value(top_gap["distribution_gap_count"]) if top_gap is not None else "-", "sub": f"{top_gap['brand']} | {top_gap['retailer']}" if top_gap is not None else "No major gap"},
                {"label": "Total Gap Exposure", "value": format_metric_value(dist["distribution_gap_count"].sum()) if len(dist) and "distribution_gap_count" in dist.columns else "-", "sub": "Store opportunities across gaps"},
                {"label": "Gap Combinations", "value": str(len(dist)), "sub": "Brand-retailer whitespace pairs"},
            ])
            c1, c2 = st.columns(2)
            with c1:
                if len(dist):
                    temp = dist.copy()
                    temp["gap_label"] = temp["brand"].astype(str) + " | " + temp["retailer"].astype(str)
                    fig = bar_chart(temp, "gap_label", "distribution_gap_count", "Largest Brand-Retailer Whitespace", top_n=12)
                    if fig:
                        chart_panel(fig)
            with c2:
                if len(dist):
                    retailer_gap = dist.groupby("retailer", dropna=False)["distribution_gap_count"].sum().reset_index().sort_values("distribution_gap_count", ascending=False)
                    fig = bar_chart(retailer_gap, "retailer", "distribution_gap_count", "Retailer Exposure to Distribution Gaps", top_n=10)
                    if fig:
                        chart_panel(fig)
            headline = "Whitespace remains concentrated in a limited set of brand-retailer combinations."
            what = f"Largest whitespace is {safe_top_value(dist, 'distribution_gap_count', ascending=False, label_cols=['brand', 'retailer'])}."
            why = "Distribution gaps are often the fastest route to incremental sales when velocity fundamentals are already healthy."
            action = "Prioritize sell-in on the highest-gap combinations supported by strong velocity or positive momentum."
            render_insight_card("Distribution Gaps", headline, what, why, action)
            cols = [c for c in [
                "brand", "category", "retailer", "current_store_count",
                "retailer_store_universe", "distribution_gap_count", "distribution_gap_index"
            ] if c in dist.columns]
            st.dataframe(dist.sort_values("distribution_gap_count", ascending=False)[cols], use_container_width=True, hide_index=True)

        with tabs[4]:
            st.markdown("<div class='section-title'>YoY Growth</div><div class='small-note'>Separate growth leaders from declining items and understand which categories are carrying the business.</div>", unsafe_allow_html=True)
            if len(yoy):
                yoy_clean = yoy.dropna(subset=["yoy_sales_growth_pct"]).copy()
                top_yoy = best_row(yoy_clean, "yoy_sales_growth_pct", ascending=False)
                low_yoy = best_row(yoy_clean, "yoy_sales_growth_pct", ascending=True)
                render_kpi_strip([
                    {"label": "Top YoY Winner", "value": format_metric_value(top_yoy["yoy_sales_growth_pct"], "pct") if top_yoy is not None else "-", "sub": f"{top_yoy['brand']} | {top_yoy['sku_id']}" if top_yoy is not None else "No YoY leader"},
                    {"label": "Biggest Decliner", "value": format_metric_value(low_yoy["yoy_sales_growth_pct"], "pct") if low_yoy is not None else "-", "sub": f"{low_yoy['brand']} | {low_yoy['sku_id']}" if low_yoy is not None else "No decline"},
                    {"label": "Average YoY", "value": format_metric_value(yoy_clean['yoy_sales_growth_pct'].mean(), "pct") if len(yoy_clean) else "-", "sub": "Portfolio growth average"},
                ])
                c1, c2 = st.columns(2)
                with c1:
                    temp = yoy_clean.sort_values("yoy_sales_growth_pct", ascending=False).head(12).copy()
                    if len(temp):
                        temp["sku_label"] = temp["sku_id"].astype(str)
                        fig = bar_chart(temp, "sku_label", "yoy_sales_growth_pct", "Top YoY Winners", top_n=12)
                        if fig:
                            chart_panel(fig)
                with c2:
                    temp = yoy_clean.sort_values("yoy_sales_growth_pct", ascending=True).head(12).copy()
                    if len(temp):
                        temp["sku_label"] = temp["sku_id"].astype(str)
                        fig = bar_chart(temp, "sku_label", "yoy_sales_growth_pct", "Most Material YoY Decliners", top_n=12, ascending=True)
                        if fig:
                            chart_panel(fig)
                headline = "Growth is uneven, with a small set of winners offset by notable declines."
                what = f"Top YoY winner is {safe_top_value(yoy_clean, 'yoy_sales_growth_pct', ascending=False, label_cols=['brand', 'sku_id'])}."
                why = "A spread between winners and decliners usually points to assortment, distribution, or execution differences rather than category-wide pressure alone."
                action = "Scale support behind winners and diagnose pricing, facings, or placement for the steepest decliners."
                render_insight_card("Year-over-Year Growth", headline, what, why, action)
                cols = [c for c in yoy.columns if c in [
                    "sku_id", "brand", "category", "yoy_sales_growth_pct", "yoy_units_growth_pct", "exception_flags"
                ] or c.startswith("sales_") or c.startswith("units_")]
                st.dataframe(yoy.sort_values("yoy_sales_growth_pct", ascending=False, na_position="last")[cols], use_container_width=True, hide_index=True)
            else:
                st.info("Not enough history for YoY analysis.")

        with tabs[5]:
            st.markdown("<div class='section-title'>Momentum</div><div class='small-note'>Read near-term directionality to distinguish durable acceleration from weakening trends.</div>", unsafe_allow_html=True)
            if len(momentum):
                top_mom = best_row(momentum, "momentum_ratio", ascending=False)
                render_kpi_strip([
                    {"label": "Best Momentum", "value": format_metric_value(top_mom["momentum_ratio"]) if top_mom is not None else "-", "sub": f"{top_mom['brand']} | {top_mom['sku_id']}" if top_mom is not None else "No leader"},
                    {"label": "Trending Up", "value": str(int((momentum["momentum_flag"] == "Trending Up").sum())) if 'momentum_flag' in momentum.columns else "-", "sub": "Positive momentum signals"},
                    {"label": "Trending Down", "value": str(int((momentum["momentum_flag"] == "Trending Down").sum())) if 'momentum_flag' in momentum.columns else "-", "sub": "Watch-list items"},
                ])
                c1, c2 = st.columns(2)
                with c1:
                    temp = momentum.sort_values("momentum_ratio", ascending=False).head(15).copy()
                    temp["sku_label"] = temp["sku_id"].astype(str)
                    fig = bar_chart(temp, "sku_label", "momentum_ratio", "Strongest Momentum Signals", color="momentum_flag", top_n=15)
                    if fig:
                        chart_panel(fig)
                with c2:
                    mix = momentum.groupby("momentum_flag", dropna=False).size().reset_index(name="count")
                    fig = donut_chart(mix, "momentum_flag", "count", "Momentum Signal Mix")
                    if fig:
                        chart_panel(fig)
                headline = "Momentum signals show where recent velocity is accelerating or cooling."
                what = f"Best momentum signal is {safe_top_value(momentum, 'momentum_ratio', ascending=False, label_cols=['brand', 'sku_id'])}."
                why = "Momentum often gives an earlier directional read than longer-term growth metrics."
                action = "Lean into items trending up and review the drivers behind items moving down before distribution expands."
                render_insight_card("Momentum", headline, what, why, action)
                cols = [c for c in [
                    "sku_id", "brand", "category", "velocity_13w", "velocity_52w", "momentum_ratio", "momentum_flag"
                ] if c in momentum.columns]
                st.dataframe(momentum.sort_values("momentum_ratio", ascending=False, na_position="last")[cols], use_container_width=True, hide_index=True)
            else:
                st.info("Momentum could not be calculated.")

        with tabs[6]:
            st.markdown("<div class='section-title'>Recent Declines</div><div class='small-note'>Highlight short-term erosion and identify which items need immediate review.</div>", unsafe_allow_html=True)
            if len(declines):
                decline_row = best_row(declines, "wow_change_pct", ascending=True)
                render_kpi_strip([
                    {"label": "Largest WoW Decline", "value": format_metric_value(decline_row["wow_change_pct"], "pct") if decline_row is not None else "-", "sub": f"{decline_row['brand']} | {decline_row['sku_id']}" if decline_row is not None else "No decline"},
                    {"label": "Decline Alerts", "value": str(len(declines)), "sub": "Current flagged SKU-weeks"},
                    {"label": "Avg Decline", "value": format_metric_value(declines['wow_change_pct'].mean(), "pct") if len(declines) and 'wow_change_pct' in declines.columns else "-", "sub": "Across current watch list"},
                ])
                c1, c2 = st.columns(2)
                with c1:
                    temp = declines.copy()
                    temp["sku_label"] = temp["sku_id"].astype(str)
                    fig = bar_chart(temp, "sku_label", "wow_change_pct", "Most Severe Week-over-Week Declines", top_n=12, ascending=True)
                    if fig:
                        chart_panel(fig)
                with c2:
                    top_decline_sku = decline_row["sku_id"] if decline_row is not None else None
                    sales_base = normalize_columns(sales_history).copy()
                    sales_base["sku_id"] = sales_base["sku_id"].astype(str).str.strip()
                    sales_base["week_end_date"] = pd.to_datetime(sales_base["week_end_date"], errors="coerce")
                    sales_base["sales_dollars"] = pd.to_numeric(sales_base.get("sales_dollars", 0), errors="coerce").fillna(0) if "sales_dollars" in sales_base.columns else 0
                    sales_base = sales_base[sales_base["sku_id"] == str(top_decline_sku)]
                    sku_trend = sales_base.groupby("week_end_date", dropna=False).agg(weekly_sales=("sales_dollars", "sum")).reset_index().dropna()
                    fig = line_chart(sku_trend, "week_end_date", "weekly_sales", f"Weekly Trend for Most Pressured SKU: {top_decline_sku}") if top_decline_sku is not None else None
                    if fig:
                        chart_panel(fig)
                headline = "A focused set of items is showing sharp recent deterioration."
                what = f"Biggest decline alert is {safe_top_value(declines, 'wow_change_pct', ascending=True, label_cols=['brand', 'sku_id'])}."
                why = "Rapid week-over-week declines can signal out-of-stocks, lost placement, or local pricing disruption."
                action = "Validate supply, shelf conditions, and promotional timing on the largest decline items immediately."
                render_insight_card("Recent Declines", headline, what, why, action)
                cols = [c for c in [
                    "sku_id", "brand", "category", "week_end_date", "weekly_sales", "prev_week_sales", "wow_change_pct"
                ] if c in declines.columns]
                st.dataframe(declines[cols], use_container_width=True, hide_index=True)
            else:
                st.info("No recent declines detected.")

        with tabs[7]:
            st.markdown("<div class='section-title'>Shelf Productivity</div><div class='small-note'>Evaluate whether facings and shelf space are translating into productive sales output.</div>", unsafe_allow_html=True)
            if len(shelf_df):
                shelf_top = best_row(shelf_df, "space_efficiency_index", ascending=False)
                render_kpi_strip([
                    {"label": "Top SEI", "value": format_metric_value(shelf_top["space_efficiency_index"]) if shelf_top is not None else "-", "sub": f"{shelf_top['retailer']} | {shelf_top['sku_id']}" if shelf_top is not None else "No SEI leader"},
                    {"label": "Avg Sales / Facing", "value": format_metric_value(shelf_df['sales_per_facing'].mean()) if len(shelf_df) and 'sales_per_facing' in shelf_df.columns else "-", "sub": "Productivity benchmark"},
                    {"label": "Increase Facing Calls", "value": str(int((shelf_df["shelf_action"] == "Increase Facings").sum())) if 'shelf_action' in shelf_df.columns else "-", "sub": "Potential shelf growth actions"},
                ])
                c1, c2 = st.columns(2)
                with c1:
                    temp = shelf_df.sort_values("space_efficiency_index", ascending=False).head(12).copy()
                    temp["sku_label"] = temp["sku_id"].astype(str)
                    fig = bar_chart(temp, "sku_label", "space_efficiency_index", "Highest Shelf Productivity Opportunities", color="shelf_action", top_n=12)
                    if fig:
                        chart_panel(fig)
                with c2:
                    fig = scatter_chart(shelf_df, "facings", "total_sales", "Facings vs. Sales Productivity", color="retailer", hover_name="sku_id", size="space_efficiency_index")
                    if fig:
                        chart_panel(fig)
                headline = "Shelf productivity highlights where space can work harder."
                what = f"Strongest opportunity is {safe_top_value(shelf_df, 'space_efficiency_index', ascending=False, label_cols=['retailer', 'sku_id'])}."
                why = "Items outperforming category productivity benchmarks may justify more space, while laggards may require rebalancing."
                action = "Review facings allocation and reset priorities based on SEI and current sales-per-facing performance."
                render_insight_card("Shelf Productivity", headline, what, why, action)
                cols = [c for c in [
                    "store_id", "retailer", "sku_id", "brand", "category", "facings",
                    "total_sales", "total_units", "shelf_productivity_score", "sales_per_facing",
                    "category_avg_sales_per_facing", "space_efficiency_index", "shelf_action", "exception_flags"
                ] if c in shelf_df.columns]
                st.dataframe(shelf_df.sort_values("space_efficiency_index", ascending=False)[cols], use_container_width=True, hide_index=True)
            else:
                st.info("No shelf file was uploaded, so shelf productivity and SEI were not calculated.")
        st.divider()

        st.markdown("<div class='download-panel'>", unsafe_allow_html=True)
        st.markdown("<div class='section-title'>Executive Exports</div>", unsafe_allow_html=True)
        st.markdown("<div class='small-note'>Download the full analytical workbook or the polished executive PDF report. Logo branding remains embedded in both outputs.</div>", unsafe_allow_html=True)
        d1, d2 = st.columns(2)
        with d1:
            st.download_button(
                "Download Full Results Workbook",
                full_results_file,
                file_name="shelfiq_911_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        with d2:
            st.download_button(
                "Download Executive PDF Report",
                pdf_file,
                file_name="shelfiq_911_executive_report.pdf",
                mime="application/pdf",
                use_container_width=True
            )
        st.markdown("</div>", unsafe_allow_html=True)

    except Exception as e:
        st.error(f"Analysis failed: {e}")